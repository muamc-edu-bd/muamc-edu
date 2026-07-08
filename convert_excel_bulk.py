"""
convert_excel_bulk.py
=====================
Converts "Online Registration Final_list.xlsx" into the format required
by the Humanities Subject Bulk Input feature on student-list.html.

Expected output format (3 columns):
  Roll | Main Subject Code | Optional

  - Roll: plain integer (e.g. 1, 2, 3 ...)
  - Main Subject Code: all subject codes joined by "/" (e.g. 174/175/176/177/275/178/179)
  - Optional: optional subject codes joined by "/" (e.g. 265/266)

The script:
  1. Reads the source Excel file.
  2. Skips header/separator rows (non-numeric roll numbers).
  3. Skips rows with empty roll numbers.
  4. Cleans up whitespace and newlines inside cells.
  5. Joins all codes with "/" as required by the bulk input parser.
  6. Writes "Humanities_Bulk_Input_Ready.xlsx" with clean, flat data.
"""

import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SOURCE = "Online Registration Final_list.xlsx"
OUTPUT = "Humanities_Bulk_Input_Ready.xlsx"


def parse_codes(cell_value) -> str:
    """Extract all numeric codes from a cell and join with '/'."""
    if cell_value is None:
        return ""
    raw = str(cell_value).strip()
    if not raw:
        return ""
    # Extract every sequence of digits
    codes = re.findall(r"\d+", raw)
    return "/".join(codes)


def is_numeric_roll(roll_str: str) -> bool:
    """Return True if the roll string is a plain integer (possibly zero-padded)."""
    cleaned = roll_str.strip().replace("\n", "").strip()
    return cleaned.isdigit()


def main():
    # ── Load source ──────────────────────────────────────────────────────────
    wb_src = openpyxl.load_workbook(SOURCE)
    ws_src = wb_src.active

    rows_out = []  # list of (roll_str, main_codes, opt_codes)
    skipped_rows = []

    for row_idx, row in enumerate(ws_src.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue  # Skip original header

        roll_raw = str(row[0] or "").strip().replace("\n", "").strip()
        main_raw = row[1]
        opt_raw  = row[2]

        # Skip separator / sub-header rows
        if not roll_raw or not is_numeric_roll(roll_raw):
            skipped_rows.append((row_idx, roll_raw))
            continue

        # Zero-pad roll to 3 digits for consistency with DB (optional, adjust as needed)
        roll_int = int(roll_raw)
        roll_str = str(roll_int)  # plain integer, e.g. "1", "26", "103"

        main_codes = parse_codes(main_raw)
        opt_codes  = parse_codes(opt_raw)

        rows_out.append((roll_str, main_codes, opt_codes))

    # ── Build output workbook ─────────────────────────────────────────────────
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Humanities Bulk Input"

    # Styles
    header_font    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="1A3C5E")
    header_align   = Alignment(horizontal="center", vertical="center", wrap_text=False)
    data_align_c   = Alignment(horizontal="center", vertical="center")
    data_align_l   = Alignment(horizontal="left",   vertical="center")
    thin_border    = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alt_fill       = PatternFill("solid", fgColor="EBF2FB")

    # Header row
    headers = ["Roll", "Main Subject Code", "Optional"]
    col_widths = [10, 40, 20]
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws_out.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
        ws_out.column_dimensions[get_column_letter(col)].width = w

    ws_out.row_dimensions[1].height = 22

    # Data rows
    for row_num, (roll, main, opt) in enumerate(rows_out, start=2):
        fill = alt_fill if row_num % 2 == 0 else None

        c_roll = ws_out.cell(row=row_num, column=1, value=roll)
        c_main = ws_out.cell(row=row_num, column=2, value=main)
        c_opt  = ws_out.cell(row=row_num, column=3, value=opt)

        for cell, align in [(c_roll, data_align_c), (c_main, data_align_l), (c_opt, data_align_c)]:
            cell.alignment = align
            cell.border    = thin_border
            cell.font      = Font(name="Calibri", size=10)
            if fill:
                cell.fill = fill

    # Freeze header row
    ws_out.freeze_panes = "A2"

    # Auto-filter
    ws_out.auto_filter.ref = f"A1:C{len(rows_out) + 1}"

    # ── Save ─────────────────────────────────────────────────────────────────
    wb_out.save(OUTPUT)

    print(f"[DONE]")
    print(f"   Source rows read   : {ws_src.max_row - 1}")
    print(f"   Skipped (non-roll) : {len(skipped_rows)}")
    print(f"   Data rows written  : {len(rows_out)}")
    print(f"   Output file        : {OUTPUT}")
    print()
    if skipped_rows:
        print("   Skipped rows (row index -> roll value):")
        for r_idx, r_val in skipped_rows:
            print(f"     Row {r_idx:4d}: '{r_val}'")
    print()
    print("   Sample output rows:")
    for r, m, o in rows_out[:5]:
        print(f"     Roll={r:>4s} | Main={m} | Opt={o}")
    if len(rows_out) > 5:
        print(f"     ... ({len(rows_out) - 5} more rows)")


if __name__ == "__main__":
    main()
