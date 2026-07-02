"""
HSC Academic Management System — Python Flask Backend
Moinuddin Adarsha Mohila College, Sylhet
Uses PostgreSQL Database with SQLAlchemy ORM

Run:
    pip install -r requirements.txt
    python app.py

The server starts at http://localhost:5000
Open the HTML file served at http://localhost:5000/
"""

import os
import uuid
import base64
import secrets
import functools
from datetime import datetime
from werkzeug.utils import secure_filename

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass   # python-dotenv optional; fallback to defaults below

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from flask import Flask, request, jsonify, send_from_directory, abort, session, Response, redirect
from flask_cors import CORS

from models import db, Student, Mark, Teacher, Setting, Archive, PromotionLog

# ─────────────────────────────────────────────
# App Setup
# ─────────────────────────────────────────────
app = Flask(__name__, static_folder='static')
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Database Configuration
# Priority: DATABASE_URL env var → local SQLite fallback
# DATABASE_URL is loaded from .env if python-dotenv is installed.
_db_url = os.environ.get('DATABASE_URL') or f"sqlite:///{os.path.join(BASE_DIR, 'hsc_academy.db')}"
if _db_url.startswith('postgres://'):
    _db_url = _db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# pool_size / max_overflow are not supported by SQLite — only set for PostgreSQL
if _db_url.startswith('postgresql'):
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_size': 10,
        'max_overflow': 20,
        'pool_recycle': 3600,
        'pool_pre_ping': True,
    }

# Secret key for server-side session (session cookie auth)
app.secret_key = os.environ.get('SECRET_KEY') or secrets.token_hex(32)
CORS(app, supports_credentials=True)

# Initialize database
db.init_app(app)

# ─────────────────────────────────────────────
# Auto-create tables + run safe column migrations at startup.
# db.create_all() only creates MISSING tables — it never alters existing ones.
# We run ALTER TABLE ... ADD COLUMN IF NOT EXISTS for columns added after the
# initial deployment. IF NOT EXISTS makes every statement safe to re-run.
# ─────────────────────────────────────────────
_COLUMN_MIGRATIONS = [
    # (table, column, column_definition)
    ("students", "optional_subjects", "VARCHAR(50) DEFAULT ''"),
    ("students", "session",           "VARCHAR(50) DEFAULT ''"),
    ("students", "year",              "VARCHAR(10) DEFAULT ''"),
    ("students", "photo",             "TEXT DEFAULT ''"),
    ("students", "section",           "VARCHAR(50) DEFAULT ''"),
    ("students", "father",            "VARCHAR(255) DEFAULT ''"),
    ("students", "mother",            "VARCHAR(255) DEFAULT ''"),
    ("students", "dob",               "VARCHAR(50) DEFAULT ''"),
    ("students", "phone",             "VARCHAR(20) DEFAULT ''"),
    ("students", "religion",          "VARCHAR(50) DEFAULT ''"),
    ("students", "reg",               "VARCHAR(50) DEFAULT ''"),
    ("teachers", "empid",             "VARCHAR(50) DEFAULT ''"),
    ("teachers", "joining",           "VARCHAR(50) DEFAULT ''"),
    ("teachers", "address",           "TEXT DEFAULT ''"),
    ("marks",    "selected_optional", "VARCHAR(50) DEFAULT ''"),
    ("students", "student_submitted", "BOOLEAN DEFAULT FALSE"),
    ("students", "photo_base64",      "TEXT DEFAULT ''"),
    ("archive",  "photo_base64",      "TEXT DEFAULT ''"),
    ("marks",    "absent",            "BOOLEAN DEFAULT FALSE"),
]

with app.app_context():
    from sqlalchemy import text as _text
    try:
        db.create_all()
        print("[DB] Tables verified/created successfully.")
        print(f"[DB] Connected to: ...@{_db_url.split('@')[-1] if '@' in _db_url else _db_url}")
    except Exception as _e:
        print(f"[DB] WARNING: Could not create tables: {_e}")

    # Run safe column migrations
    if _db_url.startswith('postgresql'):
        # Check if migrations are already complete to avoid exclusive table locks at startup
        # Check whether ALL migration columns already exist.
        # If any column is missing, run the full migration block.
        _migration_needed = False
        try:
            with db.engine.connect() as _conn:
                for _tbl, _col, _ in _COLUMN_MIGRATIONS:
                    _res = _conn.execute(_text(
                        f"SELECT EXISTS (SELECT 1 FROM information_schema.columns "
                        f"WHERE table_name='{_tbl}' AND column_name='{_col}');"
                    )).scalar()
                    if not _res:
                        _migration_needed = True
                        print(f"[DB] Missing column detected: {_tbl}.{_col} — migrations will run.")
                        break
        except Exception:
            _migration_needed = True

        if _migration_needed:
            with db.engine.connect() as _conn:
                # Drop global unique roll constraint to allow non-globally unique roll numbers
                try:
                    _conn.execute(_text("ALTER TABLE students DROP CONSTRAINT IF EXISTS students_roll_key;"))
                    _conn.commit()
                    print("[DB] Dropped global unique roll constraint successfully.")
                except Exception as _ce:
                    print(f"[DB] Drop unique constraint skipped: {_ce}")
                    try:
                        _conn.rollback()
                    except Exception:
                        pass

                for _tbl, _col, _col_def in _COLUMN_MIGRATIONS:
                    try:
                        _conn.execute(_text(
                            f"ALTER TABLE {_tbl} ADD COLUMN IF NOT EXISTS {_col} {_col_def};"
                        ))
                        _conn.commit()
                        print(f"[DB] Column migration OK: {_tbl}.{_col}")
                    except Exception as _me:
                        print(f"[DB] Column migration skipped ({_tbl}.{_col}): {_me}")
                        try:
                            _conn.rollback()
                        except Exception:
                            pass

                # Ensure photo column is TEXT type in both students and archive tables
                for _tbl in ["students", "archive"]:
                    try:
                        _conn.execute(_text(f"ALTER TABLE {_tbl} ALTER COLUMN photo TYPE TEXT;"))
                        _conn.commit()
                        print(f"[DB] Altered photo column to TEXT type OK: {_tbl}")
                    except Exception as _pe:
                        print(f"[DB] Alter photo column to TEXT skipped ({_tbl}): {_pe}")
                        try:
                            _conn.rollback()
                        except Exception:
                            pass
        else:
            print("[DB] Schema up to date. Skipping startup DDL migrations.")

# ─────────────────────────────────────────────
# Credentials — read from .env; never hard-coded
# Create a .env file:  ADMIN_UID=admin  ADMIN_PW=yourPassword
# ─────────────────────────────────────────────
ADMIN_UID = os.environ.get('ADMIN_UID', 'admin')
ADMIN_PW  = os.environ.get('ADMIN_PW',  'admin1234')

# Credentials for Marks Entry only user
MARKS_UID = os.environ.get('MARKS_UID', 'teacher')
MARKS_PW  = os.environ.get('MARKS_PW',  'teacher123')

# ─────────────────────────────────────────────
# Auth decorator — protects all /api/* routes
# ─────────────────────────────────────────────
def require_auth(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('authenticated'):
            return jsonify({'ok': False, 'message': 'Unauthorized. Please log in.'}), 401
        return f(*args, **kwargs)
    return decorated

def require_admin(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('authenticated') or session.get('role') != 'admin':
            return jsonify({'ok': False, 'message': 'Forbidden. Admin access required.'}), 403
        return f(*args, **kwargs)
    return decorated

# ─────────────────────────────────────────────
# File paths (for photo storage only)
# ─────────────────────────────────────────────
PHOTOS_DIR = os.path.join(BASE_DIR, 'photos')
os.makedirs(PHOTOS_DIR, exist_ok=True)


# Photos stored directly in the database as base64 data URLs to prevent loss
# on ephemeral hosting environments (e.g. Render, Heroku).
# ─────────────────────────────────────────────
def _save_photo_file(student_id: str, data_url: str) -> str:
    """
    Save photo data URL directly to database (and optionally backup to file).
    Returns the relative URL path to store in the DB.
    """
    if not data_url:
        return ''
    if not data_url.startswith('data:'):
        return data_url
    try:
        header, encoded = data_url.split(',', 1)
        mime = header.split(';')[0].split(':')[1]
        ext_map = {'image/jpeg': '.jpg', 'image/png': '.png',
                   'image/gif': '.gif', 'image/webp': '.webp'}
        ext = ext_map.get(mime, '.jpg')
        filename = secure_filename(f'{student_id}{ext}')
        filepath = os.path.join(PHOTOS_DIR, filename)
        with open(filepath, 'wb') as f:
            f.write(base64.b64decode(encoded))
        return f'/photos/{filename}'
    except Exception:
        pass
    return ''


def _delete_photo_file(photo_url: str):
    if not photo_url or not photo_url.startswith('/photos/'):
        return
    filename = secure_filename(os.path.basename(photo_url))
    filepath = os.path.join(PHOTOS_DIR, filename)
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
    except OSError:
        pass


# ─────────────────────────────────────────────
# Serve static files
# ─────────────────────────────────────────────
@app.route('/photos/<path:filename>')
def serve_photo(filename):
    filepath = os.path.join(PHOTOS_DIR, filename)
    if not os.path.exists(filepath):
        # File missing on disk (e.g., ephemeral hosting container restart).
        # Lazy-restore from DB photo_base64
        student_id = filename.split('.')[0]
        student = Student.query.filter_by(id=student_id).first()
        b64_data = ''
        if student and student.photo_base64:
            b64_data = student.photo_base64
        else:
            archived = Archive.query.filter_by(id=student_id).first()
            if archived and archived.photo_base64:
                b64_data = archived.photo_base64
        
        if b64_data and b64_data.startswith('data:'):
            try:
                header, encoded = b64_data.split(',', 1)
                os.makedirs(PHOTOS_DIR, exist_ok=True)
                with open(filepath, 'wb') as f:
                    f.write(base64.b64decode(encoded))
            except Exception:
                pass

    return send_from_directory(PHOTOS_DIR, filename)


# ─────────────────────────────────────────────
# Serve HTML pages
# Use send_from_directory only — prevents path traversal
# ─────────────────────────────────────────────
def _serve_html(name, inject_api=True):
    html_path = os.path.join(BASE_DIR, name)
    if not os.path.exists(html_path):
        abort(404, description=f'{name} not found')
    with open(html_path, 'r', encoding='utf-8') as f:
        html = f.read()

    # Inject Kalpurush Font globally for all served HTML pages
    font_inject = (
        '<link href="https://fonts.maateen.me/kalpurush/font.css" rel="stylesheet">\n'
        '<style> * { font-family: "Kalpurush", sans-serif !important; } </style>\n'
    )
    html = html.replace('</head>', font_inject + '</head>', 1)

    if inject_api:
        inject = '<script>window.USE_API = true; window.API_BASE = "";</script>'
        html = html.replace('</head>', inject + '\n</head>', 1)
    return Response(html, mimetype='text/html')


@app.route('/')
def index():
    if not session.get('authenticated'):
        return redirect('/login')
    return _serve_html('index.html')

@app.route('/login')
def login_page():
    return _serve_html('login.html', inject_api=False)

@app.route('/result')
def result_page():
    return _serve_html('result.html')

@app.route('/marks-entry.html')
def marks_entry_page():
    return _serve_html('marks-entry.html')

@app.route('/marks-input.html')
def marks_input_page():
    return _serve_html('marks-input.html')

@app.route('/index.html')
def index_html_page():
    if not session.get('authenticated'):
        return redirect('/login')
    return _serve_html('index.html')

@app.route('/teacher-input.html')
def teacher_input_page():
    if not session.get('authenticated'):
        return redirect('/login')
    return _serve_html('teacher-input.html')

# ── Extracted standalone pages (formerly inline tabs) ────────────────────────

@app.route('/Teachers.html')
def teachers_page():
    """Standalone Teachers management page."""
    if not session.get('authenticated'):
        return redirect('/login')
    return _serve_html('Teachers.html')

@app.route('/Result-card.html')
def result_card_page():
    """Standalone Result Card page."""
    if not session.get('authenticated'):
        return redirect('/login')
    return _serve_html('Result-card.html')

@app.route('/Admit-card.html')
def admit_card_page():
    """Standalone Admit Card page."""
    if not session.get('authenticated'):
        return redirect('/login')
    return _serve_html('Admit-card.html')

@app.route('/bulk-photo.html')
def bulk_photo_page():
    """Bulk Photo Upload page."""
    if not session.get('authenticated'):
        return redirect('/login')
    return _serve_html('bulk-photo.html')

@app.route('/analytics')
def analytics_page():
    return _serve_html('analytics.html')

@app.route('/result_analytics.html')
def result_analytics_page():
    """Result Analytics — tabulation sheet generator page."""
    if not session.get('authenticated'):
        return redirect('/login')
    return _serve_html('result_analytics.html')

@app.route('/result_summery.html')
def result_summery_page():
    """Public result summary page — accessible via QR code scan, no auth required."""
    return _serve_html('result_summery.html', inject_api=False)

@app.route('/student-portal')
def student_portal_page():
    """Public student self-service portal — no authentication required."""
    return _serve_html('student-portal.html', inject_api=False)

@app.route('/<path:page>')
def html_pages(page):
    # Only serve .html files from BASE_DIR; prevent directory traversal
    if not page.endswith('.html') or '/' in page or '..' in page:
        abort(404)
    # Auth-gate any .html page that is not the login or student-portal page
    if page not in ('login.html', 'student-portal.html', 'result_summery.html') and not session.get('authenticated'):
        return redirect('/login')
    # Use _serve_html to ensure the global font and API scripts are correctly injected
    return _serve_html(page, inject_api=(page not in ('login.html', 'student-portal.html', 'result_summery.html')))


# ─────────────────────────────────────────────
# RESULT ANALYTICS — Tabulation Sheet API
# ─────────────────────────────────────────────

@app.route('/api/tabulation-sheet/filters', methods=['GET'])
@require_auth
def tabulation_filters():
    """
    Return distinct sessions, exam_types and subject list for a cls + group.
    Used to populate cascading filter dropdowns on the Result Analytics page.
    Query params: cls, group
    """
    cls_val   = (request.args.get('cls')   or '').strip()
    group_val = (request.args.get('group') or '').strip()

    if not cls_val or not group_val:
        return jsonify({'ok': False, 'message': 'cls and group are required'}), 400

    import copy as _copy

    # Distinct sessions from matching students
    students = Student.query.filter_by(cls=cls_val, group=group_val).all()
    sessions_set = set()
    for s in students:
        val = (s.session or s.year or '').strip()
        if val:
            sessions_set.add(val)
    sessions = sorted(sessions_set, reverse=True)

    # Distinct exam_types from marks of these students
    student_ids = [s.id for s in students]
    exam_types_set = set()
    if student_ids:
        marks_rows = Mark.query.filter(Mark.student_id.in_(student_ids)).all()
        for m in marks_rows:
            if m.exam_type:
                exam_types_set.add(m.exam_type)
    exam_types = sorted(exam_types_set)

    # Subject list filtered by paper (same logic used everywhere in codebase)
    raw_subs = _copy.deepcopy(SUBJECT_LIST.get(group_val, []))
    if cls_val == 'Class-XI':
        raw_subs = [s for s in raw_subs if not ('2nd' in s['name'] or '2nd Paper' in s['name'] or s['name'].endswith(' 2nd'))]
    elif cls_val == 'Class-XII':
        raw_subs = [s for s in raw_subs if not ('1st' in s['name'] or '1st Paper' in s['name'] or s['name'].endswith(' 1st'))]
    temp = []
    for sub in raw_subs:
        s = _copy.deepcopy(sub)
        if s['code'] in ['116', '267']:
            suffix = ' 1st Paper' if cls_val == 'Class-XI' else ' 2nd Paper'
            s['name'] = s['name'].replace(' 1st Paper', '').replace(' 2nd Paper', '') + suffix
        temp.append(s)

    return jsonify({
        'ok': True,
        'sessions':   sessions,
        'exam_types': exam_types,
        'subjects':   [{'name': s['name'], 'code': s['code'], 'hasPrac': s.get('hasPrac', False)} for s in temp],
    })


@app.route('/api/tabulation-sheet', methods=['GET'])
@require_auth
def get_tabulation_sheet():
    """
    Return all data required to render one tabulation sheet for a specific
    class / group / session / exam_type / subject_code combination.

    Query params:
        cls          — e.g. 'Class-XI'
        group        — 'Science' | 'Humanities' | 'Business'
        session      — e.g. '2024-2025'
        exam_type    — e.g. 'First Terminal'
        subject_code — e.g. '101'
    """
    cls_val      = (request.args.get('cls')          or '').strip()
    group_val    = (request.args.get('group')         or '').strip()
    session_val  = (request.args.get('session')       or '').strip().replace('\u2013', '-').replace('\u2014', '-')
    exam_type    = (request.args.get('exam_type')     or '').strip()
    subject_code = (request.args.get('subject_code')  or '').strip()

    if not all([cls_val, group_val, session_val, exam_type, subject_code]):
        return jsonify({'ok': False,
                        'message': 'cls, group, session, exam_type and subject_code are all required'}), 400

    import copy as _copy

    # Resolve subject metadata from the canonical SUBJECT_LIST
    raw_subs = _copy.deepcopy(SUBJECT_LIST.get(group_val, []))
    if cls_val == 'Class-XI':
        raw_subs = [s for s in raw_subs if not ('2nd' in s['name'] or '2nd Paper' in s['name'] or s['name'].endswith(' 2nd'))]
    elif cls_val == 'Class-XII':
        raw_subs = [s for s in raw_subs if not ('1st' in s['name'] or '1st Paper' in s['name'] or s['name'].endswith(' 1st'))]
    temp = []
    for sub in raw_subs:
        s = _copy.deepcopy(sub)
        if s['code'] in ['116', '267']:
            suffix = ' 1st Paper' if cls_val == 'Class-XI' else ' 2nd Paper'
            s['name'] = s['name'].replace(' 1st Paper', '').replace(' 2nd Paper', '') + suffix
        temp.append(s)

    subject_meta = next((s for s in temp if s['code'] == subject_code), None)
    if subject_meta is None:
        return jsonify({'ok': False,
                        'message': f'Subject code {subject_code} not found for {group_val} / {cls_val}'}), 404

    # Fetch students sorted by roll number
    students = Student.query.filter_by(cls=cls_val, group=group_val).filter(
        (Student.session == session_val) | (Student.year == session_val)
    ).all()

    # Sort numerically where possible, then alphabetically
    def _roll_key(s):
        try:
            return (0, int(s.roll))
        except (ValueError, TypeError):
            return (1, s.roll or '')

    students = sorted(students, key=_roll_key)

    if not students:
        college_setting = Setting.query.filter_by(key='collegeName').first()
        return jsonify({
            'ok': True,
            'rows': [],
            'summary': {'total_students': 0, 'total_appeared': 0, 'passed': 0, 'failed': 0, 'absent': 0},
            'subject': subject_meta,
            'cls': cls_val, 'group': group_val,
            'session': session_val, 'exam_type': exam_type,
            'college_name': (college_setting.value if college_setting else '') or 'Moinuddin Adarsha Mohila College, Sylhet',
        })

    student_ids = [s.id for s in students]

    # Fetch marks for this exam_type + subject_code only (efficient single query)
    mark_rows = Mark.query.filter(
        Mark.student_id.in_(student_ids),
        Mark.exam_type    == exam_type,
        Mark.subject_code == subject_code,
    ).all()
    mark_map = {m.student_id: m for m in mark_rows}

    cq_max   = subject_meta.get('cqMax',  70)
    mcq_max  = subject_meta.get('mcqMax', 30)
    has_prac = bool(subject_meta.get('hasPrac'))

    rows = []
    total_appeared = passed = failed = absent_count = 0

    for idx, stu in enumerate(students, 1):
        m = mark_map.get(stu.id)

        # Absent if: no mark row, explicit absent flag, or cq+mcq both empty
        if m is None:
            is_absent = True
        elif m.absent:
            is_absent = True
        elif str(m.cq or '') == '' and str(m.mcq or '') == '':
            is_absent = True
        else:
            is_absent = False

        if is_absent:
            absent_count += 1
            rows.append({
                'sl': idx, 'roll': stu.roll, 'name': stu.name,
                'cq': None, 'mcq': None, 'prac': None,
                'total': None, 'gpa': None, 'grade': 'Ab', 'absent': True,
            })
            continue

        total_appeared += 1
        cq   = min(int(m.cq   or 0), cq_max)
        mcq  = min(int(m.mcq  or 0), mcq_max)
        prac = min(int(m.prac or 0), 25) if has_prac else 0
        total = cq + mcq + prac
        grade_letter, gpa_point = _grade_letter(total)

        # A subject is failed if the student gets an F total OR fails any
        # individual component (CQ / MCQ / Practical) below its pass mark.
        sub_passed = _subject_passed(cq, mcq, prac, has_prac, cq_max, mcq_max)
        if grade_letter == 'F' or not sub_passed:
            failed += 1
        else:
            passed += 1

        rows.append({
            'sl': idx, 'roll': stu.roll, 'name': stu.name,
            'cq': cq, 'mcq': mcq,
            'prac': prac if has_prac else None,
            'total': total, 'gpa': gpa_point, 'grade': grade_letter,
            'absent': False,
        })

    college_setting = Setting.query.filter_by(key='collegeName').first()
    college_name = (college_setting.value if college_setting else '') or \
                   'Moinuddin Adarsha Mohila College, Sylhet'

    return jsonify({
        'ok':           True,
        'college_name': college_name,
        'cls':          cls_val,
        'group':        group_val,
        'session':      session_val,
        'exam_type':    exam_type,
        'subject':      subject_meta,
        'summary': {
            'total_students': len(students),
            'total_appeared': total_appeared,
            'passed':         passed,
            'failed':         failed,
            'absent':         absent_count,
        },
        'rows': rows,
    })


# ─────────────────────────────────────────────
# HEALTH CHECK  (no auth — for deployment diagnosis)
# ─────────────────────────────────────────────
@app.route('/api/health', methods=['GET'])
def health_check():
    db_ok = False
    db_info = {}
    try:
        from sqlalchemy import text
        with db.engine.connect() as conn:
            conn.execute(text('SELECT 1'))
        student_count = Student.query.count()
        teacher_count = Teacher.query.count()
        marks_count = Mark.query.count()
        db_ok = True
        safe_url = _db_url.split('@')[-1] if '@' in _db_url else _db_url
        db_info = {
            'connected': True,
            'host': safe_url,
            'students': student_count,
            'teachers': teacher_count,
            'marks_entries': marks_count,
        }
    except Exception as e:
        db_info = {'connected': False, 'error': str(e)}

    return jsonify({
        'ok': db_ok,
        'status': 'healthy' if db_ok else 'unhealthy',
        'database': db_info,
        'version': '1.0.0',
        'timestamp': datetime.utcnow().isoformat() + 'Z',
    }), 200 if db_ok else 503


# ─────────────────────────────────────────────
# AUTH
# ─────────────────────────────────────────────
@app.route('/api/login', methods=['POST'])
def login():
    body = request.get_json(force=True, silent=True) or {}
    uid = body.get('uid', '').strip()
    pw = body.get('pw', '')

    # Admin Logic: User mentioned "one password only" for admin. 
    # We'll allow admin to login with just the password OR with the admin UID + PW.
    if pw == ADMIN_PW:
        session['authenticated'] = True
        session['role'] = 'admin'
        session.permanent = False
        return jsonify({'ok': True, 'role': 'admin', 'message': 'Admin login successful'})
    
    # Marks Entry Login: Needs both ID and Password
    if uid == MARKS_UID and pw == MARKS_PW:
        session['authenticated'] = True
        session['role'] = 'marks_entry'
        session.permanent = False
        return jsonify({'ok': True, 'role': 'marks_entry', 'message': 'Marks entry login successful'})

    return jsonify({'ok': False, 'message': 'Invalid credentials'}), 401


@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'ok': True, 'message': 'Logged out'})


@app.route('/api/auth-status', methods=['GET'])
def auth_status():
    return jsonify({
        'authenticated': session.get('authenticated', False),
        'role': session.get('role', None)
    })


# ─────────────────────────────────────────────
# STUDENT PORTAL  (public — no auth required)
# ─────────────────────────────────────────────

@app.route('/api/student-portal/lookup', methods=['GET'])
def student_portal_lookup():
    """
    Public endpoint: find a student by class + group + session + roll.
    Returns minimal info (name, submission status) — no sensitive data.
    """
    cls_val     = (request.args.get('cls') or '').strip()
    group_val   = (request.args.get('group') or '').strip()
    session_val = (request.args.get('session') or '').strip().replace('–', '-').replace('\u2013', '-')
    roll_val    = (request.args.get('roll') or '').strip()

    if not cls_val or not group_val or not session_val or not roll_val:
        return jsonify({'ok': False, 'message': 'All fields are required: cls, group, session, roll'}), 400

    student = Student.query.filter_by(
        cls=cls_val,
        group=group_val,
        roll=roll_val,
    ).filter(
        (Student.session == session_val) | (Student.year == session_val)
    ).first()

    if not student:
        return jsonify({'ok': False, 'message': 'No student found with the provided information. Please check your details and try again.'}), 404

    submitted = bool(student.student_submitted)
    return jsonify({
        'ok':        True,
        'id':        student.id,
        'name':      student.name,
        'cls':       student.cls,
        'group':     student.group,
        'session':   student.session or student.year,
        'roll':      student.roll,
        'submitted': submitted,
        # If already submitted, show current values so student can view them
        'photo':              student.photo if submitted else None,
        'optionalSubjects':   student.optional_subjects,
    })


@app.route('/api/student-portal/submit/<sid>', methods=['POST'])
def student_portal_submit(sid):
    """
    Public endpoint: student uploads photo + selects optional subject.
    One-time only — sets student_submitted = True after first save.
    """
    student = Student.query.filter_by(id=sid).first()
    if not student:
        return jsonify({'ok': False, 'message': 'Student not found'}), 404

    if student.student_submitted:
        return jsonify({
            'ok':      False,
            'message': 'You have already submitted your information. Please contact the college administration to make changes.',
            'already_submitted': True,
        }), 409

    body             = request.get_json(force=True, silent=True) or {}
    optional_subject = (body.get('optional_subject') or '').strip()
    photo_data       = body.get('photo', '')

    if not optional_subject and not photo_data:
        return jsonify({'ok': False, 'message': 'Please upload a photo and/or select an optional subject.'}), 400

    # Save photo if provided
    if photo_data and photo_data.startswith('data:'):
        _delete_photo_file(student.photo)
        student.photo_base64 = photo_data
        new_photo = _save_photo_file(student.id, photo_data)
        if new_photo:
            student.photo = new_photo

    # Save optional subject if provided
    if optional_subject:
        student.optional_subjects = optional_subject

    # Lock: prevent further student edits
    student.student_submitted = True

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Database error: {str(e)}'}), 500

    return jsonify({
        'ok':      True,
        'message': 'Your information has been saved successfully. You cannot make further changes.',
        'photo':   student.photo,
        'optionalSubjects': student.optional_subjects,
    })


# ─────────────────────────────────────────────
# STUDENTS  (/api/students)
# ─────────────────────────────────────────────
@app.route('/api/students', methods=['GET'])
@require_auth
def get_students():
    cls_filter   = request.args.get('cls')
    group_filter = request.args.get('group')
    session_val  = request.args.get('session')
    if session_val:
        session_val = session_val.strip().replace('–', '-').replace('\u2013', '-')
    q            = (request.args.get('q') or '').lower()

    query = Student.query

    if cls_filter:
        query = query.filter_by(cls=cls_filter)
    if group_filter:
        query = query.filter_by(group=group_filter)
    if session_val:
        query = query.filter(
            (Student.session == session_val) | (Student.year == session_val)
        )
    if q:
        query = query.filter(
            (Student.name.ilike(f'%{q}%')) |
            (Student.roll.ilike(f'%{q}%'))
        )

    try:
        students = [s.to_dict() for s in query.all()]
        return jsonify({'ok': True, 'data': students})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'message': f'Database error: {str(e)}'}), 500


@app.route('/api/students', methods=['POST'])
@require_auth
def add_student():
    body = request.get_json(force=True, silent=True) or {}
    required = ['name', 'roll', 'cls', 'group']
    for field in required:
        if not body.get(field):
            return jsonify({'ok': False, 'message': f'Field "{field}" is required'}), 400

    # Check if roll already exists
    if Student.query.filter_by(roll=body['roll']).first():
        return jsonify({'ok': False, 'message': 'A student with this roll number already exists'}), 409

    student_id   = str(uuid.uuid4().hex[:16])
    photo_b64    = body.get('photo', '')
    photo_url    = _save_photo_file(student_id, photo_b64)

    student = Student(
        id=student_id,
        name=body.get('name', ''),
        roll=body.get('roll', ''),
        reg=body.get('reg', ''),
        cls=body.get('cls', ''),
        group=body.get('group', ''),
        section=body.get('section', ''),
        father=body.get('father', ''),
        mother=body.get('mother', ''),
        dob=body.get('dob', ''),
        phone=body.get('phone', ''),
        religion=body.get('religion', ''),
        year=(body.get('year') or '').strip().replace('–', '-').replace('\u2013', '-'),
        session=(body.get('session') or body.get('year') or '').strip().replace('–', '-').replace('\u2013', '-'),
        photo=photo_url,
        photo_base64=photo_b64,
        optional_subjects=body.get('optional_subject', ''),
    )
    db.session.add(student)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Database error: {str(e)}'}), 500
    return jsonify({'ok': True, 'data': student.to_dict()}), 201


@app.route('/api/students/<sid>', methods=['GET'])
@require_auth
def get_student(sid):
    student = Student.query.filter_by(id=sid).first()
    if not student:
        return jsonify({'ok': False, 'message': 'Student not found'}), 404
    return jsonify({'ok': True, 'data': student.to_dict()})


def _resolve_optional_subjects(group: str, optional_subject: str) -> list:
    """
    Return the resolved subject list for a student, substituting the generic
    'optional' placeholder entries with the student's specific chosen codes.

    optional_subject format: "CODE1/CODE2"  e.g. "178/179" or "265/266"
    """
    import copy
    opt_codes = [c.strip() for c in (optional_subject or '').split('/') if c.strip()]
    base = copy.deepcopy(SUBJECT_LIST.get(group, []))

    if not opt_codes:
        # No optional specified — return base list as-is (caller handles placeholders)
        return base

    # Code→name lookup (extend as needed)
    OPT_NAMES = {
        '178': 'Biology 1st Paper',           '179': 'Biology 2nd Paper',
        '265': 'Higher Mathematics 1st Paper','266': 'Higher Mathematics 2nd Paper',
        '121': 'Logic 1st Paper',             '122': 'Logic 2nd Paper',
        '273': 'Home Science 1st Paper',      '274': 'Home Science 2nd Paper',
        '109': 'Economics 1st Paper',         '110': 'Economics 2nd Paper',
    }

    resolved = []
    for sub in base:
        if not sub.get('optional'):
            resolved.append(sub)
            continue
        # All groups take both optional subjects (one compulsory elective, one optional).
        resolved.append({**sub, 'name': OPT_NAMES.get(sub['code'], sub['name'])})
    return resolved


@app.route('/api/students/<sid>/subjects', methods=['GET'])
@require_auth
def get_student_subjects(sid):
    """Return the resolved subject list for a student, respecting their optional_subject."""
    student = Student.query.filter_by(id=sid).first()
    if not student:
        return jsonify({'ok': False, 'message': 'Student not found'}), 404
    subjects = _resolve_optional_subjects(student.group, getattr(student, 'optional_subjects', '') or '')
    cls = student.cls or 'Class-XI'
    if cls == 'Class-XI':
        subjects = [sub for sub in subjects if not ('2nd' in sub['name'] or '2nd Paper' in sub['name'] or sub['name'].endswith(' 2nd'))]
    elif cls == 'Class-XII':
        subjects = [sub for sub in subjects if not ('1st' in sub['name'] or '1st Paper' in sub['name'] or sub['name'].endswith(' 1st'))]

    # Dynamically rename Sociology (116) and Islamic History & Culture (267) to append paper number
    import copy
    temp_subs = []
    for sub in subjects:
        s = copy.deepcopy(sub)
        if s['code'] in ['116', '267']:
            suffix = ' 1st Paper' if cls == 'Class-XI' else ' 2nd Paper'
            s['name'] = s['name'].replace(' 1st Paper', '').replace(' 2nd Paper', '') + suffix
        temp_subs.append(s)
    subjects = temp_subs

    return jsonify({'ok': True, 'data': subjects, 'optional_subject': getattr(student, 'optional_subjects', '')})


@app.route('/api/students/<sid>', methods=['PUT'])
@require_admin
def update_student(sid):
    student = Student.query.filter_by(id=sid).first()
    if not student:
        return jsonify({'ok': False, 'message': 'Student not found'}), 404

    body    = request.get_json(force=True, silent=True) or {}
    allowed = ['name', 'roll', 'reg', 'cls', 'group', 'section', 'father', 'mother',
               'dob', 'phone', 'religion', 'year', 'session', 'photo', 'optional_subjects']

    # Also accept the 'optional_subject' alias (without the trailing 's') that the frontend sends
    if 'optional_subject' in body and 'optional_subjects' not in body:
        body['optional_subjects'] = body['optional_subject']

    for key in allowed:
        if key not in body:
            continue
        if key == 'photo':
            val = body[key]
            if val and val.startswith('data:'):
                _delete_photo_file(student.photo)
                student.photo_base64 = val
                student.photo = _save_photo_file(sid, val)
            elif val == '':
                _delete_photo_file(student.photo)
                student.photo_base64 = ''
                student.photo = ''
        else:
            val = body[key]
            if key in ('session', 'year') and isinstance(val, str):
                val = val.strip().replace('–', '-').replace('\u2013', '-')
            setattr(student, key, val)

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Database error: {str(e)}'}), 500
    return jsonify({'ok': True, 'data': student.to_dict()})


@app.route('/api/students/<sid>', methods=['DELETE'])
@require_admin
def delete_student(sid):
    student = Student.query.filter_by(id=sid).first()
    if not student:
        return jsonify({'ok': False, 'message': 'Student not found'}), 404

    _delete_photo_file(student.photo)
    # cascade='all, delete-orphan' on the relationship automatically removes marks
    db.session.delete(student)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Database error: {str(e)}'}), 500
    return jsonify({'ok': True, 'message': 'Student deleted'})


# ─────────────────────────────────────────────
# STUDENT IMPORT
# Photo column removed — upload photos individually per student
# ─────────────────────────────────────────────
def _parse_excel_file(file):
    if not HAS_OPENPYXL:
        return None, "openpyxl not installed. Run: pip install openpyxl"
    try:
        wb = load_workbook(file)
        ws = wb.active
        headers = []
        for cell in ws[1]:
            headers.append(cell.value.lower().replace(' ', '').strip() if cell.value else '')

        column_map = {
            'name': 'name', 'roll': 'roll', 'registration': 'reg',
            'class': 'cls', 'group': 'group', 'section': 'section',
            'father': 'father', 'mother': 'mother', 'dob': 'dob',
            'phone': 'phone', 'religion': 'religion', 'session': 'session',
            'year': 'year', 'optionalsubject': 'optional_subject',
            'optional': 'optional_subject',
        }
        col_indices = {}
        for expected, field in column_map.items():
            for i, header in enumerate(headers):
                if expected in header:
                    col_indices[field] = i
                    break

        students, preview_samples = [], []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(cell is None for cell in row):
                continue
            student = {}
            for field, col_idx in col_indices.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    student[field] = str(row[col_idx]).strip()
            if student.get('name') and student.get('roll'):
                students.append(student)
                if len(preview_samples) < 3:
                    preview_samples.append(student)
        return students, preview_samples
    except Exception as e:
        return None, f"Failed to parse Excel: {str(e)}"


@app.route('/api/students/import/preview', methods=['POST'])
@require_auth
def import_preview():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'message': 'No file provided'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'ok': False, 'message': 'No file selected'}), 400
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'ok': False, 'message': 'Only Excel files (.xlsx, .xls) are supported'}), 400
    students, preview = _parse_excel_file(file)
    if students is None:
        return jsonify({'ok': False, 'message': preview}), 400
    return jsonify({'ok': True, 'data': students,
                    'preview': {'rows': len(students), 'samples': preview}})


@app.route('/api/students/import', methods=['POST'])
@app.route('/api/students/bulk-import', methods=['POST'])  # Alias for frontend consistency
@require_auth
def import_students():
    body        = request.get_json(force=True, silent=True) or {}
    import_list = body.get('students', [])
    if not import_list:
        return jsonify({'ok': False, 'message': 'No students provided'}), 400

    existing_rolls = {s.roll for s in Student.query.all()}
    imported = skipped = 0

    for student_data in import_list:
        if not student_data.get('name') or not student_data.get('roll'):
            skipped += 1
            continue
        if student_data['roll'] in existing_rolls:
            skipped += 1
            continue

        # FIX: use uuid to guarantee unique IDs even in rapid bulk imports
        student = Student(
            id=uuid.uuid4().hex[:16],
            name=student_data.get('name', ''),
            roll=student_data.get('roll', ''),
            reg=student_data.get('reg', ''),
            cls=student_data.get('cls', 'Class-XI'),
            group=student_data.get('group', 'Science'),
            section=student_data.get('section', ''),
            father=student_data.get('father', ''),
            mother=student_data.get('mother', ''),
            dob=student_data.get('dob', ''),
            phone=student_data.get('phone', ''),
            religion=student_data.get('religion', ''),
            year=student_data.get('year', '').strip().replace('–', '-').replace('\u2013', '-'),
            session=student_data.get('session', student_data.get('year', '')).strip().replace('–', '-').replace('\u2013', '-'),
            photo='',
            optional_subjects=(
                student_data.get('optional_subject') or
                student_data.get('optional_subjects') or
                student_data.get('optionalSubjects') or
                ''
            ),
        )
        db.session.add(student)
        existing_rolls.add(student_data['roll'])
        imported += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Database error during import: {str(e)}'}), 500
    return jsonify({'ok': True, 'imported': imported, 'skipped': skipped,
                    'message': f'Successfully imported {imported} students'})


# ─────────────────────────────────────────────
# TEACHERS  (/api/teachers)
# ─────────────────────────────────────────────
@app.route('/api/teachers', methods=['GET'])
@require_auth
def get_teachers():
    q     = (request.args.get('q') or '').lower()
    query = Teacher.query

    if q:
        query = query.filter(
            (Teacher.name.ilike(f'%{q}%')) |
            (Teacher.email.ilike(f'%{q}%')) |
            (Teacher.subject.ilike(f'%{q}%'))
        )

    teachers = [t.to_dict() for t in query.all()]
    return jsonify({'ok': True, 'data': teachers})


@app.route('/api/teachers', methods=['POST'])
@require_admin
def add_teacher():
    body     = request.get_json(force=True, silent=True) or {}
    required = ['name', 'email', 'phone', 'subject', 'qualification']
    for field in required:
        if not body.get(field):
            return jsonify({'ok': False, 'message': f'Field "{field}" is required'}), 400

    teacher = Teacher(
        id='T' + str(int(datetime.utcnow().timestamp() * 1000)),
        name=body.get('name', ''),
        email=body.get('email', ''),
        phone=body.get('phone', ''),
        subject=body.get('subject', ''),
        classes=body.get('classes', '—'),
        qualification=body.get('qualification', ''),
        experience=int(body.get('experience', 0) or 0),
        empid=body.get('empid', ''),
        joining=body.get('joining', ''),
        address=body.get('address', ''),
    )
    db.session.add(teacher)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Database error: {str(e)}'}), 500
    return jsonify({'ok': True, 'data': teacher.to_dict()}), 201


@app.route('/api/teachers/<tid>', methods=['PUT'])
@require_admin
def update_teacher(tid):
    teacher = Teacher.query.filter_by(id=tid).first()
    if not teacher:
        return jsonify({'ok': False, 'message': 'Teacher not found'}), 404

    body    = request.get_json(force=True, silent=True) or {}
    allowed = ['name', 'email', 'phone', 'subject', 'classes', 'qualification',
               'experience', 'empid', 'joining', 'address']
    for key in allowed:
        if key in body:
            setattr(teacher, key, body[key])

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Database error: {str(e)}'}), 500
    return jsonify({'ok': True, 'data': teacher.to_dict()})


@app.route('/api/teachers/<tid>', methods=['DELETE'])
@require_admin
def delete_teacher(tid):
    teacher = Teacher.query.filter_by(id=tid).first()
    if not teacher:
        return jsonify({'ok': False, 'message': 'Teacher not found'}), 404

    db.session.delete(teacher)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Database error: {str(e)}'}), 500
    return jsonify({'ok': True, 'message': 'Teacher deleted'})


# ─────────────────────────────────────────────
# MARKS  (/api/marks)
# ─────────────────────────────────────────────
def _get_marks_dict(student_id=None):
    """Convert database marks to the hierarchical JSON format used by frontend.
    Format: {student_id: {exam_type: {subject_code: {cq,mcq,prac,ey,examType,year},
                                      selectedOptional: str}}}
    """
    result = {}

    if student_id:
        if isinstance(student_id, (list, tuple, set)):
            marks = Mark.query.filter(Mark.student_id.in_(student_id)).all()
        else:
            marks = Mark.query.filter_by(student_id=student_id).all()
    else:
        marks = Mark.query.all()

    for mark in marks:
        sid  = mark.student_id
        exam = mark.exam_type
        subj = mark.subject_code

        result.setdefault(sid, {}).setdefault(exam, {})

        result[sid][exam][subj] = {
            'cq':       mark.cq,
            'mcq':      mark.mcq,
            'prac':     mark.prac,
            'absent':   mark.absent or False,
            'ey':       mark.year,
            'examType': mark.exam_type,
            'year':     mark.year,
        }

        if mark.selected_optional:
            result[sid][exam]['selectedOptional'] = mark.selected_optional

    return result


@app.route('/api/marks', methods=['GET'])
@require_auth
def get_all_marks():
    return jsonify({'ok': True, 'data': _get_marks_dict()})


@app.route('/api/marks', methods=['DELETE'])
@require_admin
def clear_all_marks():
    Mark.query.delete()
    db.session.commit()
    return jsonify({'ok': True, 'message': 'All marks cleared'})


@app.route('/api/marks/<sid>', methods=['GET'])
@require_auth
def get_marks(sid):
    result = {}
    for mark in Mark.query.filter_by(student_id=sid).all():
        exam = mark.exam_type
        subj = mark.subject_code
        result.setdefault(exam, {})
        result[exam][subj] = {
            'cq':       mark.cq,
            'mcq':      mark.mcq,
            'prac':     mark.prac,
            'absent':   mark.absent or False,
            'ey':       mark.year,
            'examType': mark.exam_type,
            'year':     mark.year,
        }
        if mark.selected_optional:
            result[exam]['selectedOptional'] = mark.selected_optional

    return jsonify({'ok': True, 'data': result})


@app.route('/api/students/bulk-photo', methods=['POST'])
@require_auth
def bulk_photo_upload():
    """
    Bulk photo upload endpoint.
    Receives a list of { roll, photo } entries, matches each roll to a student
    (filtered by cls + group + optional session), saves photo using _save_photo_file.

    Request body:
    {
      "cls":     "Class-XI",
      "group":   "Science",
      "session": "2024-2025",   // optional
      "entries": [
        { "roll": "5",  "photo": "data:image/jpeg;base64,..." },
        { "roll": "12", "photo": "data:image/jpeg;base64,..." }
      ]
    }
    """
    body    = request.get_json(force=True, silent=True) or {}
    cls     = (body.get('cls')     or '').strip()
    group   = (body.get('group')   or '').strip()
    ses     = (body.get('session') or '').strip().replace('\u2013', '-').replace('\u2014', '-')
    entries = body.get('entries', [])

    if not cls or not group:
        return jsonify({'ok': False, 'message': 'cls and group are required'}), 400
    if not entries:
        return jsonify({'ok': False, 'message': 'No entries provided'}), 400

    results    = []
    saved_cnt  = 0
    notfnd_cnt = 0

    for entry in entries:
        roll      = str(entry.get('roll') or '').strip()
        photo_b64 = entry.get('photo', '')

        if not roll:
            results.append({'roll': roll, 'name': None, 'status': 'invalid', 'message': 'No roll number'})
            continue

        # Find student: cls + group, optionally + session
        query = Student.query.filter_by(cls=cls, group=group)
        if ses:
            query = query.filter(
                (Student.session == ses) | (Student.year == ses)
            )
        students = query.all()
        student = None
        for s in students:
            r1 = str(s.roll or '').strip().lstrip('0')
            r2 = str(roll or '').strip().lstrip('0')
            if r1 == r2:
                student = s
                break
            if (s.roll or '').strip().isdigit() and roll.isdigit():
                if int(s.roll) == int(roll):
                    student = s
                    break

        if not student:
            results.append({'roll': roll, 'name': None, 'status': 'not_found',
                            'message': f'No student with roll {roll} in {cls} / {group}'})
            notfnd_cnt += 1
            continue

        if not photo_b64 or not photo_b64.startswith('data:'):
            results.append({'roll': roll, 'name': student.name, 'status': 'error',
                            'message': 'Invalid photo data'})
            continue

        had_photo = bool(student.photo)
        _delete_photo_file(student.photo)                          # delete old photo file
        new_url = _save_photo_file(student.id, photo_b64)         # save new photo file
        if not new_url:
            results.append({'roll': roll, 'name': student.name, 'status': 'error',
                            'message': 'Failed to save photo file'})
            continue

        student.photo_base64 = photo_b64
        student.photo = new_url
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            results.append({'roll': roll, 'name': student.name, 'status': 'error',
                            'message': f'DB error: {str(e)}'})
            continue

        status = 'replaced' if had_photo else 'saved'
        saved_cnt += 1
        results.append({
            'roll':     roll,
            'name':     student.name,
            'status':   status,
            'message':  'Photo replaced' if had_photo else 'Photo saved',
            'photoUrl': new_url,
        })

    return jsonify({
        'ok':        True,
        'results':   results,
        'saved':     saved_cnt,
        'not_found': notfnd_cnt,
        'total':     len(entries),
    })


@app.route('/api/public/result-summary', methods=['GET'])
def public_result_summary():
    """
    Public endpoint — no auth required. Called by result_summery.html (QR code page).
    Returns full result data for a student given sid + exam type.
    """
    sid       = (request.args.get('sid') or '').strip()
    exam_type = (request.args.get('exam') or '').strip()

    if not sid:
        return jsonify({'ok': False, 'message': 'Missing sid parameter'}), 400

    student = Student.query.filter_by(id=sid).first()
    if not student:
        return jsonify({'ok': False, 'message': 'Student not found'}), 404

    # Get all marks for this student
    marks_data = _get_marks_dict(student.id)
    stu_marks_all = marks_data.get(student.id, {})

    # Resolve exam type — use provided or latest
    if exam_type and exam_type in stu_marks_all:
        resolved_exam = exam_type
    elif stu_marks_all:
        resolved_exam = list(stu_marks_all.keys())[-1]
    else:
        return jsonify({'ok': False, 'message': 'No marks found for this student'}), 404

    stu_marks = stu_marks_all[resolved_exam]

    # Resolve subjects for this student's group+class
    import copy
    cls = student.cls or 'Class-XI'
    group = student.group or 'Science'
    optional_subject = getattr(student, 'optional_subjects', '') or ''
    subs = _resolve_optional_subjects(group, optional_subject)
    if cls == 'Class-XI':
        subs = [s for s in subs if not ('2nd' in s['name'] or '2nd Paper' in s['name'] or s['name'].endswith(' 2nd'))]
    elif cls == 'Class-XII':
        subs = [s for s in subs if not ('1st' in s['name'] or '1st Paper' in s['name'] or s['name'].endswith(' 1st'))]
    temp_subs = []
    for sub in subs:
        s = copy.deepcopy(sub)
        if s['code'] in ['116', '267']:
            suffix = ' 1st Paper' if cls == 'Class-XI' else ' 2nd Paper'
            s['name'] = s['name'].replace(' 1st Paper', '').replace(' 2nd Paper', '') + suffix
        temp_subs.append(s)
    subs = temp_subs

    # Resolve optional codes
    selected_optional = stu_marks.get('selectedOptional', '')
    opt_codes = [c.strip() for c in (optional_subject or '').split('/') if c.strip()]
    if not opt_codes and selected_optional:
        LEGACY_MAP = {
            'scienceBio': ['178', '179'], 'scienceMath': ['265', '266'],
            'humLogic': ['121', '122'],   'humHome': ['273', '274'],
            'busEcon': ['109', '110'],    'busHome': ['273', '274'],
        }
        opt_codes = LEGACY_MAP.get(selected_optional, [])

    # Compute per-subject marks and totals
    subject_rows = []
    total_gpa = cnt = fail_cnt = total_mark = 0
    has_absent = False
    has_fail = False

    # Compute highest marks across all peers in same class+group
    peers = Student.query.filter_by(cls=cls, group=group).all()
    peer_ids = [p.id for p in peers]
    all_peer_marks = _get_marks_dict(peer_ids)
    highest_marks = {}
    for sub in subs:
        max_mark = 0
        for p in peers:
            pm = (all_peer_marks.get(p.id) or {}).get(resolved_exam, {})
            mm = pm.get(sub['code'], {})
            pcq  = min(int(mm.get('cq')  or 0), sub.get('cqMax', 70))
            pmcq = min(int(mm.get('mcq') or 0), sub.get('mcqMax', 30))
            pprac = min(int(mm.get('prac') or 0), 25) if sub.get('hasPrac') else 0
            ptot = pcq + pmcq + pprac
            if ptot > max_mark:
                max_mark = ptot
        highest_marks[sub['code']] = max_mark

    for sub in subs:
        is_optional = sub.get('optional', False) and (sub['code'] in opt_codes)
        m = stu_marks.get(sub['code'], {})
        # Absent = no mark row, explicit absent flag, or both cq/mcq empty
        absent = (not m) or m.get('absent', False) or \
                 (str(m.get('cq', '')) == '' and str(m.get('mcq', '')) == '')
        if absent:
            if not is_optional:
                has_absent = True
            cq = mcq = prac = tot = 0
            lg, gp = 'Ab', 0.0
            sub_passed = True  # absent — does not count as component-level fail
        else:
            cq   = min(int(m.get('cq')  or 0), sub.get('cqMax', 70))
            mcq  = min(int(m.get('mcq') or 0), sub.get('mcqMax', 30))
            prac = min(int(m.get('prac') or 0), 25) if sub.get('hasPrac') else 0
            tot  = cq + mcq + prac
            lg, gp = _grade_letter(tot)
            # Check component-level pass marks (applies to optional subjects too,
            # but optional subject failure does NOT cause overall failure)
            sub_passed = _subject_passed(
                cq, mcq, prac,
                bool(sub.get('hasPrac')),
                sub.get('cqMax', 70),
                sub.get('mcqMax', 30),
            )
            if not is_optional:
                if lg == 'F' or not sub_passed:
                    has_fail = True

        if not absent:
            if is_optional:
                if gp > 2.0:
                    total_gpa += (gp - 2.0)
            else:
                total_gpa += gp
                cnt += 1
                if lg == 'F' or not sub_passed:
                    fail_cnt += 1   # F grade OR component-level fail
            total_mark += tot

        subject_rows.append({
            'name':        sub['name'],
            'code':        sub['code'],
            'cq':          '' if absent else cq,
            'mcq':         '' if absent else mcq,
            'prac':        ('' if absent else prac) if sub.get('hasPrac') else None,
            'total':       '' if absent else tot,
            'highest':     highest_marks.get(sub['code'], 0),
            'gpa':         round(gp, 2),
            'lg':          lg,
            'is_optional': is_optional,
            'has_prac':    bool(sub.get('hasPrac')),
            'absent':      absent,
            'sub_passed':  None if absent else bool(sub_passed),
        })

    # GPA = 0 if any compulsory subject failed; absent subjects excluded from GPA
    avg = 0.0 if (fail_cnt > 0 or has_absent or has_fail) else (min(round(total_gpa / cnt, 2), 5.0) if cnt else 0.0)
    # passed: no F-grade failures AND appeared in at least 1 subject
    passed = (fail_cnt == 0 and cnt > 0 and not has_absent and not has_fail)
    # mMax counts only non-absent subjects
    mMax = sum(
        (s.get('cqMax', 70) + s.get('mcqMax', 30) + (25 if s.get('hasPrac') else 0))
        for s in subs
        if not (stu_marks.get(s['code'], {}) and
                (stu_marks.get(s['code'], {}).get('absent', False) or
                 (str(stu_marks.get(s['code'], {}).get('cq', '')) == '' and
                  str(stu_marks.get(s['code'], {}).get('mcq', '')) == '')))
        and stu_marks.get(s['code'])
    )
    if mMax == 0:  # fallback for display if all absent
        mMax = sum((s.get('cqMax', 70) + s.get('mcqMax', 30) + (25 if s.get('hasPrac') else 0)) for s in subs)

    # Merit position
    peer_scores = []
    for p in peers:
        pt, pg, _ = _compute_student_result(p.id, all_peer_marks, p.group, getattr(p, 'optional_subjects', '') or '', p.cls)
        peer_scores.append((p.id, pg, pt))
    peer_scores.sort(key=lambda x: (-x[1], -x[2]))
    merit_pos   = next((i + 1 for i, ps in enumerate(peer_scores) if ps[0] == student.id), None)
    merit_total = len(peer_scores)

    # Resolve optional subject label
    OPTIONAL_SUBJECTS_LABELS = {
        'Science':     [{'id':'scienceBio','label':'Biology (1st & 2nd Paper)'},{'id':'scienceMath','label':'Higher Mathematics (1st & 2nd Paper)'}],
        'Humanities':  [{'id':'humLogic',  'label':'Logic (1st & 2nd Paper)'},  {'id':'humHome',   'label':'Home Science (1st & 2nd Paper)'}],
        'Business':    [{'id':'busEcon',   'label':'Economics (1st & 2nd Paper)'},{'id':'busHome',  'label':'Home Science (1st & 2nd Paper)'}],
    }
    opt_label = 'N/A'
    if selected_optional:
        for o in OPTIONAL_SUBJECTS_LABELS.get(group, []):
            if o['id'] == selected_optional:
                opt_label = o['label']
                break

    ey = ''
    for v in stu_marks.values():
        if isinstance(v, dict) and v.get('ey'):
            ey = v['ey']
            break
    if not ey:
        ey = student.year or student.session or ''

    return jsonify({
        'ok': True,
        'student': student.to_dict(),
        'exam_type': resolved_exam,
        'exam_year': ey,
        'subject_rows': subject_rows,
        'total_mark': total_mark,
        'max_mark': mMax,
        'gpa': avg,
        'passed': passed,
        'fail_count': fail_cnt,
        'merit_position': merit_pos,
        'merit_total': merit_total,
        'optional_label': opt_label,
    })


@app.route('/api/public/result', methods=['GET'])
def public_result_search():
    cls  = request.args.get('cls')
    grp  = request.args.get('group')
    roll = request.args.get('roll')

    if not cls or not grp or not roll:
        return jsonify({'ok': False, 'message': 'Missing search parameters'}), 400

    student = Student.query.filter_by(cls=cls, group=grp, roll=roll).first()
    if not student:
        return jsonify({'ok': False, 'message': 'Student not found with provided details.'}), 404

    # Get marks for this student
    marks_data = _get_marks_dict(student.id)

    # Compute merit position (rank within same cls + group)
    peers = Student.query.filter_by(cls=student.cls, group=student.group).all()
    peer_ids = [p.id for p in peers]
    all_marks = _get_marks_dict(peer_ids)
    peer_scores = []
    for peer in peers:
        t, g, _ = _compute_student_result(peer.id, all_marks, peer.group, getattr(peer, "optional_subjects", "") or "")
        peer_scores.append((peer.id, g, t))
    peer_scores.sort(key=lambda x: (-x[1], -x[2]))
    merit_position = next((i + 1 for i, ps in enumerate(peer_scores) if ps[0] == student.id), None)

    return jsonify({
        'ok': True,
        'student': student.to_dict(),
        'marks': marks_data.get(student.id, {}),
        'merit_position': merit_position,
        'total_peers': len(peers),
    })


@app.route('/api/marks/<sid>', methods=['POST'])
@require_auth
def save_marks(sid):
    body          = request.get_json(force=True, silent=True) or {}
    exam_type     = body.get('examType')
    subject_marks = body.get('marks', {})

    if not exam_type:
        return jsonify({'ok': False, 'message': 'examType is required'}), 400

    if not Student.query.filter_by(id=sid).first():
        return jsonify({'ok': False, 'message': 'Student not found'}), 404

    # Delete existing marks for this student and exam type
    Mark.query.filter_by(student_id=sid, exam_type=exam_type).delete()

    selected_optional = subject_marks.get('selectedOptional', '')

    for subject_code, marks_data in subject_marks.items():
        if subject_code == 'selectedOptional':
            continue

        # FIX: Validate and safely convert mark values
        try:
            # If subject is marked absent, store zeros and flag
            is_absent = bool(marks_data.get('absent', False))
            cq   = 0 if is_absent else int(float(marks_data.get('cq')  or 0))
            mcq  = 0 if is_absent else int(float(marks_data.get('mcq') or 0))
            prac = 0 if is_absent else int(float(marks_data.get('prac') or 0))
        except (ValueError, TypeError):
            return jsonify({'ok': False, 'message': f'Invalid mark format for subject {subject_code}. Marks must be numeric.'}), 400

        mark = Mark(
            student_id=sid,
            exam_type=exam_type,
            year=str(marks_data.get('year') or body.get('year') or '').strip().replace('–', '-').replace('\u2013', '-'),
            subject_code=subject_code,
            cq=cq,
            mcq=mcq,
            prac=prac,
            absent=is_absent,
            selected_optional=selected_optional,
        )
        db.session.add(mark)

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Database error: {str(e)}'}), 500
    
    return jsonify({'ok': True, 'message': 'Marks saved successfully'})


@app.route('/api/marks/batch', methods=['GET', 'POST'])
@require_auth
def get_batch_marks():
    # Support both GET (legacy, short lists) and POST (large student lists).
    # GET with 250+ student IDs produces a ~4800-char URL which exceeds server
    # limits (414 Too Long). POST sends ids in the request body to avoid this.
    if request.method == 'POST':
        body = request.get_json(force=True, silent=True) or {}
        ids_str = ','.join(body.get('ids', []))
        year_filter = (body.get('year') or '').strip().replace('\u2013', '-').replace('\u2014', '-')
    else:
        ids_str = request.args.get('ids', '')
        year_filter = (request.args.get('year') or '').strip().replace('\u2013', '-').replace('\u2014', '-')

    if not ids_str:
        return jsonify({'ok': True, 'data': {}})
    sids = [sid.strip() for sid in ids_str.split(',') if sid.strip()]
    if not sids:
        return jsonify({'ok': True, 'data': {}})

    marks = _get_marks_dict(sids)

    # If a specific year/session is requested, filter marks to only that year.
    # This prevents marks from different sessions (e.g. 2024-2025 vs 2025-2026)
    # from colliding and overwriting each other in the dict.
    if year_filter:
        filtered = {}
        for sid, exams in marks.items():
            filtered_exams = {}
            for exam_type, subjects in exams.items():
                filtered_subjects = {}
                for key, val in subjects.items():
                    if key == 'selectedOptional':
                        filtered_subjects[key] = val
                    elif isinstance(val, dict) and val.get('year') == year_filter:
                        filtered_subjects[key] = val
                if filtered_subjects:
                    filtered_exams[exam_type] = filtered_subjects
            if filtered_exams:
                filtered[sid] = filtered_exams
        marks = filtered

    return jsonify({'ok': True, 'data': marks})

@app.route('/api/marks/batch-subject', methods=['POST'])
@require_auth
def save_batch_subject_marks():

    body = request.get_json(force=True, silent=True) or {}
    subject_code = body.get('subjectCode')
    exam_type = body.get('examType')
    year = body.get('year')
    entries = body.get('entries', [])

    if not subject_code or not exam_type or not year:
        return jsonify({'ok': False, 'message': 'subjectCode, examType, and year are required'}), 400

    try:
        for entry in entries:
            sid = entry.get('studentId')
            if not sid:
                continue

            cq_raw = entry.get('cq')
            mcq_raw = entry.get('mcq')
            prac_raw = entry.get('prac')
            is_absent = bool(entry.get('absent', False))

            # A row is empty only if not absent AND all marks are blank
            is_empty = (not is_absent) and \
                       (cq_raw == '' or cq_raw is None) and \
                       (mcq_raw == '' or mcq_raw is None) and \
                       (prac_raw == '' or prac_raw is None)

            # Delete existing marks for this student, exam type, and subject code
            Mark.query.filter_by(student_id=sid, exam_type=exam_type, subject_code=subject_code).delete()

            if not is_empty:
                try:
                    # When absent: store zeros with absent=True flag
                    cq   = 0 if is_absent else int(float(cq_raw or 0))
                    mcq  = 0 if is_absent else int(float(mcq_raw or 0))
                    prac = 0 if is_absent else int(float(prac_raw or 0))
                except (ValueError, TypeError):
                    return jsonify({'ok': False, 'message': f'Invalid mark format for student {sid}. Marks must be numeric.'}), 400

                selected_optional = entry.get('selectedOptional')
                if selected_optional is None:
                    existing = Mark.query.filter_by(student_id=sid, exam_type=exam_type).first()
                    selected_optional = existing.selected_optional if existing else ''

                mark = Mark(
                    student_id=sid,
                    exam_type=exam_type,
                    year=str(year).strip().replace('–', '-').replace('\u2013', '-'),
                    subject_code=subject_code,
                    cq=cq,
                    mcq=mcq,
                    prac=prac,
                    absent=is_absent,
                    selected_optional=selected_optional or ''
                )
                db.session.add(mark)

        db.session.commit()
        return jsonify({'ok': True, 'message': 'Batch marks saved successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Database error: {str(e)}'}), 500


@app.route('/api/marks/import', methods=['POST'])
@require_auth
def import_marks():
    """Bulk import marks from Excel rows."""
    body = request.get_json(force=True, silent=True) or {}
    rows = body.get('rows', [])

    if not isinstance(rows, list) or not rows:
        return jsonify({'ok': False, 'message': 'No mark rows provided'}), 400

    roll_to_id = {s.roll: s.id for s in Student.query.all()}
    imported = skipped = 0
    warnings = []

    for row in rows:
        if not isinstance(row, dict):
            skipped += 1
            continue

        roll      = str(row.get('Roll') or row.get('roll') or '').strip()
        exam_type = str(row.get('Exam Type') or row.get('examType') or row.get('exam') or '').strip()
        year      = str(row.get('Year') or row.get('year') or '').strip().replace('–', '-').replace('\u2013', '-')

        if not roll or not exam_type or not year:
            skipped += 1
            continue

        student_id = roll_to_id.get(roll)
        if not student_id:
            skipped += 1
            warnings.append(f'Student roll not found: {roll}')
            continue

        Mark.query.filter_by(student_id=student_id, exam_type=exam_type).delete()

        for key, value in row.items():
            if key in ['Roll', 'roll', 'Exam Type', 'examType', 'Year', 'year']:
                continue
            if value is None or value == '':
                continue

            mark_str = str(value).strip()
            parts    = [p.strip() for p in mark_str.split('/') if p.strip()]
            if len(parts) < 2:
                continue

            try:
                cq   = int(float(parts[0]))
                mcq  = int(float(parts[1]))
                prac = int(float(parts[2])) if len(parts) > 2 else 0
            except ValueError:
                continue

            mark = Mark(
                student_id=student_id,
                exam_type=exam_type,
                year=year,
                subject_code=key.strip(),
                cq=cq,
                mcq=mcq,
                prac=prac,
            )
            db.session.add(mark)

        imported += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Database error during marks import: {str(e)}'}), 500
    return jsonify({'ok': True, 'imported': imported, 'skipped': skipped,
                     'warnings': warnings,
                    'message': f'{imported} rows imported, {skipped} rows skipped'})


@app.route('/api/marks/bulk-import', methods=['POST'])
@require_auth
def bulk_import_marks():
    """Bulk import marks from the frontend Excel parser.

    Expected JSON body:
        {
            "entries": [
                {
                    "studentId": "abc123",
                    "examType":  "Annual",
                    "year":      "2024–2025",
                    "marks": {
                        "101": {"cq":"65","mcq":"25","prac":"","ey":"2024–2025","year":"2024–2025"},
                        "102": {"cq":"58","mcq":"22","prac":"","ey":"2024–2025","year":"2024–2025"},
                        "selectedOptional": "scienceBio"
                    }
                },
                ...
            ]
        }

    All entries are saved to PostgreSQL in a single transaction.
    """
    body    = request.get_json(force=True, silent=True) or {}
    entries = body.get('entries', [])

    if not isinstance(entries, list) or not entries:
        return jsonify({'ok': False, 'message': 'No entries provided'}), 400

    # Pre-validate: collect all student IDs referenced
    valid_student_ids = {s.id for s in Student.query.all()}

    imported = 0
    skipped  = 0

    for entry in entries:
        student_id = entry.get('studentId', '')
        exam_type  = entry.get('examType', '')
        year       = str(entry.get('year') or '').strip().replace('–', '-').replace('\u2013', '-')
        marks_data = entry.get('marks', {})

        if not student_id or not exam_type:
            skipped += 1
            continue

        if student_id not in valid_student_ids:
            skipped += 1
            continue

        # Delete existing marks for this student + exam type (replace strategy)
        Mark.query.filter_by(student_id=student_id, exam_type=exam_type).delete()

        selected_optional = marks_data.get('selectedOptional', '')

        for subject_code, mark_obj in marks_data.items():
            if subject_code == 'selectedOptional':
                continue
            if not isinstance(mark_obj, dict):
                continue

            try:
                cq   = int(float(mark_obj.get('cq') or 0))
                mcq  = int(float(mark_obj.get('mcq') or 0))
                prac = int(float(mark_obj.get('prac') or 0))
            except (ValueError, TypeError):
                continue   # skip invalid marks silently

            mark = Mark(
                student_id=student_id,
                exam_type=exam_type,
                year=str(mark_obj.get('year') or year).strip().replace('–', '-').replace('\u2013', '-'),
                subject_code=subject_code,
                cq=cq,
                mcq=mcq,
                prac=prac,
                selected_optional=selected_optional,
            )
            db.session.add(mark)

        imported += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False,
                        'message': f'Database error during bulk marks import: {str(e)}'}), 500

    return jsonify({'ok': True, 'imported': imported, 'skipped': skipped,
                    'message': f'{imported} student marks saved to PostgreSQL, {skipped} skipped'})


@app.route('/api/marks/<sid>/<exam_type>', methods=['DELETE'])
@require_auth
def delete_marks(sid, exam_type):
    Mark.query.filter_by(student_id=sid, exam_type=exam_type).delete()
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Database error: {str(e)}'}), 500
    return jsonify({'ok': True, 'message': 'Marks deleted'})


# ─────────────────────────────────────────────
# DASHBOARD STATS
# ─────────────────────────────────────────────
@app.route('/api/stats', methods=['GET'])
@require_auth
def get_stats():
    students = Student.query.all()
    marks    = Mark.query.all()
    teachers = Teacher.query.all()

    stats = {
        'total':       len(students),
        'classXI':     sum(1 for s in students if s.cls == 'Class-XI'),
        'classXII':    sum(1 for s in students if s.cls == 'Class-XII'),
        'science':     sum(1 for s in students if s.group == 'Science'),
        'humanities':  sum(1 for s in students if s.group == 'Humanities'),
        'business':    sum(1 for s in students if s.group == 'Business'),
        'withResults': len({m.student_id for m in marks}),
        'withPhotos':  sum(1 for s in students if s.photo),
        'teachers':    len(teachers),
    }
    return jsonify({'ok': True, 'data': stats})


# ─────────────────────────────────────────────
# SETTINGS
# ─────────────────────────────────────────────
@app.route('/api/settings', methods=['GET'])
@require_auth
def get_settings():
    settings = {s.key: s.value for s in Setting.query.all()}
    return jsonify({'ok': True, 'data': settings})


@app.route('/api/settings', methods=['POST'])
@require_auth
def save_settings():
    body = request.get_json(force=True, silent=True) or {}

    for key, value in body.items():
        setting = Setting.query.filter_by(key=key).first()
        if setting:
            setting.value = value
        else:
            db.session.add(Setting(key=key, value=value))

    db.session.commit()
    return jsonify({'ok': True, 'message': 'Settings saved'})


# ─────────────────────────────────────────────
# GRADING  helpers
# ─────────────────────────────────────────────
def _grade_letter(total):
    """Return (letter_grade, gpa_point) for a given theory total."""
    if total >= 80: return 'A+', 5.0
    if total >= 70: return 'A',  4.0
    if total >= 60: return 'A-', 3.5
    if total >= 50: return 'B',  3.0
    if total >= 40: return 'C',  2.0
    if total >= 33: return 'D',  1.0
    return 'F', 0.0


def _subject_passed(cq, mcq, prac, has_prac, cq_max=70, mcq_max=30):
    """
    Determine whether a student passed a subject based on component-level pass marks.

    Rules:
      - CQ-only subjects (mcq_max == 0, e.g. English 1st/2nd Paper, 100 marks):
            CQ pass mark = 33 out of 100
      - Subjects WITHOUT practical (has_prac=False, cqMax=70, mcqMax=30):
            CQ pass mark  = 23 out of 70
            MCQ pass mark = 10 out of 30
      - Subjects WITH practical (has_prac=True, cqMax=50, mcqMax=25, pracMax=25):
            CQ pass mark          = 17 out of 50
            MCQ pass mark         =  8 out of 25
            Practical pass mark   =  8 out of 25

    Returns True only when ALL applicable components meet their pass mark.
    A student who fails ANY component fails the subject regardless of total.
    """
    if mcq_max == 0:
        # CQ-only subject (English 1st/2nd Paper — 100 marks, no MCQ)
        return cq >= 33
    if has_prac:
        # Subject with practical: CQ(50) + MCQ(25) + Practical(25)
        return cq >= 17 and mcq >= 8 and prac >= 8
    # Subject without practical: CQ(70) + MCQ(30)
    return cq >= 23 and mcq >= 10


SUBJECT_LIST = {
    'Science': [
        {'name': 'Bangla 1st Paper',      'code': '101', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Bangla 2nd Paper',      'code': '102', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'English 1st Paper',     'code': '107', 'hasPrac': False, 'cqMax': 100, 'mcqMax': 0,  'optional': False},
        {'name': 'English 2nd Paper',     'code': '108', 'hasPrac': False, 'cqMax': 100, 'mcqMax': 0,  'optional': False},
        {'name': 'ICT',                   'code': '275', 'hasPrac': True,  'cqMax': 50,  'mcqMax': 25, 'optional': False},
        {'name': 'Physics 1st Paper',     'code': '174', 'hasPrac': True,  'cqMax': 50,  'mcqMax': 25, 'optional': False},
        {'name': 'Physics 2nd Paper',     'code': '175', 'hasPrac': True,  'cqMax': 50,  'mcqMax': 25, 'optional': False},
        {'name': 'Chemistry 1st Paper',   'code': '176', 'hasPrac': True,  'cqMax': 50,  'mcqMax': 25, 'optional': False},
        {'name': 'Chemistry 2nd Paper',   'code': '177', 'hasPrac': True,  'cqMax': 50,  'mcqMax': 25, 'optional': False},
        {'name': 'Biology 1st Paper',     'code': '178', 'hasPrac': True,  'cqMax': 50,  'mcqMax': 25, 'optional': True},
        {'name': 'Biology 2nd Paper',     'code': '179', 'hasPrac': True,  'cqMax': 50,  'mcqMax': 25, 'optional': True},
        {'name': 'Higher Math 1st Paper', 'code': '265', 'hasPrac': True,  'cqMax': 50,  'mcqMax': 25, 'optional': True},
        {'name': 'Higher Math 2nd Paper', 'code': '266', 'hasPrac': True,  'cqMax': 50,  'mcqMax': 25, 'optional': True},
    ],
    'Humanities': [
        {'name': 'Bangla 1st Paper',             'code': '101', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Bangla 2nd Paper',             'code': '102', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'English 1st Paper',            'code': '107', 'hasPrac': False, 'cqMax': 100, 'mcqMax': 0,  'optional': False},
        {'name': 'English 2nd Paper',            'code': '108', 'hasPrac': False, 'cqMax': 100, 'mcqMax': 0,  'optional': False},
        {'name': 'ICT',                          'code': '275', 'hasPrac': True,  'cqMax': 50,  'mcqMax': 25, 'optional': False},
        {'name': 'Civics & Good Governance 1st', 'code': '269', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Civics & Good Governance 2nd', 'code': '270', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Economics 1st Paper',          'code': '109', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Economics 2nd Paper',          'code': '110', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Sociology',                    'code': '116', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Social Work 1st Paper',        'code': '271', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Social Work 2nd Paper',        'code': '272', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Islamic History & Culture',    'code': '267', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Logic 1st Paper',              'code': '121', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': True},
        {'name': 'Logic 2nd Paper',              'code': '122', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': True},
        {'name': 'Home Science 1st Paper',       'code': '273', 'hasPrac': True,  'cqMax': 50,  'mcqMax': 25, 'optional': True},
        {'name': 'Home Science 2nd Paper',       'code': '274', 'hasPrac': True,  'cqMax': 50,  'mcqMax': 25, 'optional': True},
    ],
    'Business': [
        {'name': 'Bangla 1st Paper',                 'code': '101', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Bangla 2nd Paper',                 'code': '102', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'English 1st Paper',                'code': '107', 'hasPrac': False, 'cqMax': 100, 'mcqMax': 0,  'optional': False},
        {'name': 'English 2nd Paper',                'code': '108', 'hasPrac': False, 'cqMax': 100, 'mcqMax': 0,  'optional': False},
        {'name': 'ICT',                              'code': '275', 'hasPrac': True,  'cqMax': 50,  'mcqMax': 25, 'optional': False},
        {'name': 'Accounting 1st Paper',             'code': '253', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Accounting 2nd Paper',             'code': '254', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Business Org. & Mgmt 1st',         'code': '277', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Business Org. & Mgmt 2nd',         'code': '278', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Finance, Banking & Insurance 1st', 'code': '292', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Finance, Banking & Insurance 2nd', 'code': '293', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': False},
        {'name': 'Economics 1st Paper',              'code': '109', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': True},
        {'name': 'Economics 2nd Paper',              'code': '110', 'hasPrac': False, 'cqMax': 70,  'mcqMax': 30, 'optional': True},
        {'name': 'Home Science 1st Paper',           'code': '273', 'hasPrac': True,  'cqMax': 50,  'mcqMax': 25, 'optional': True},
        {'name': 'Home Science 2nd Paper',           'code': '274', 'hasPrac': True,  'cqMax': 50,  'mcqMax': 25, 'optional': True},
    ],
}


# ─────────────────────────────────────────────
# EXPORT CSV
# ─────────────────────────────────────────────
@app.route('/api/export/csv', methods=['GET'])
@require_auth
def export_csv():
    import io, csv

    students     = Student.query.all()
    marks        = _get_marks_dict()
    cls_filter   = request.args.get('cls', '')
    group_filter = request.args.get('group', '')
    exam_filter  = request.args.get('examType', '')

    if cls_filter:
        students = [s for s in students if s.cls == cls_filter]
    if group_filter:
        students = [s for s in students if s.group == group_filter]

    output = io.StringIO()
    writer = csv.writer(output)
    groups  = [group_filter] if group_filter else ['Science', 'Humanities', 'Business']
    classes = [cls_filter]   if cls_filter   else ['Class-XI', 'Class-XII']

    for g in groups:
        for c in classes:
            grp_students = [s for s in students if s.group == g and s.cls == c]
            if not grp_students:
                continue

            # Filter subjects based on class (Class-XI -> 1st Paper; Class-XII -> 2nd Paper)
            subs = SUBJECT_LIST.get(g, [])
            if c == 'Class-XI':
                subs = [sub for sub in subs if not ('2nd' in sub['name'] or '2nd Paper' in sub['name'] or sub['name'].endswith(' 2nd'))]
            elif c == 'Class-XII':
                subs = [sub for sub in subs if not ('1st' in sub['name'] or '1st Paper' in sub['name'] or sub['name'].endswith(' 1st'))]

            # Dynamically rename Sociology (116) and Islamic History & Culture (267) to append paper number
            import copy
            temp_subs = []
            for sub in subs:
                s = copy.deepcopy(sub)
                if s['code'] in ['116', '267']:
                    suffix = ' 1st Paper' if c == 'Class-XI' else ' 2nd Paper'
                    s['name'] = s['name'].replace(' 1st Paper', '').replace(' 2nd Paper', '') + suffix
                temp_subs.append(s)
            subs = temp_subs

            writer.writerow([f'{g} Group — {c}'])
            header = ['#', 'Name', 'Roll Number', 'Registration']
            for sub in subs:
                nm = sub['name']
                header += [f'{nm} (CQ)', f'{nm} (MCQ)',
                           f'{nm} (PR)' if sub['hasPrac'] else f'{nm} (—)']
            header += ['Total Marks', 'GPA', 'Result']
            writer.writerow(header)

            for i, stu in enumerate(grp_students, 1):
                sid       = stu.id
                exam_keys = list(marks.get(sid, {}).keys())
                tgt       = exam_filter or (exam_keys[-1] if exam_keys else None)
                stu_marks = marks.get(sid, {}).get(tgt, {}) if tgt else {}
                optional_selected = stu_marks.get('selectedOptional')
                opt_subject_str = getattr(stu, 'optional_subjects', '') or ''
                opt_codes = [code.strip() for code in opt_subject_str.split('/') if code.strip()]

                total_gpa = cnt = fail_cnt = total_mark = 0
                has_absent = False
                has_fail = False
                row = [i, stu.name, stu.roll, stu.reg or '']

                # Resolve optional codes
                student_opt_codes = opt_codes
                if not student_opt_codes and optional_selected:
                    LEGACY_MAP = {
                        'scienceBio': ['178', '179'],
                        'scienceMath': ['265', '266'],
                        'humLogic': ['121', '122'],
                        'humHome': ['273', '274'],
                        'busEcon': ['109', '110'],
                        'busHome': ['273', '274'],
                    }
                    student_opt_codes = LEGACY_MAP.get(optional_selected, [])

                for sub in subs:
                    is_optional = sub.get('optional', False) and (sub['code'] in student_opt_codes)

                    m      = stu_marks.get(sub['code'], {})
                    # Absent = no mark row, explicit absent flag, or both cq/mcq empty
                    absent = (not m) or m.get('absent', False) or \
                             (m.get('cq', '') == '' and m.get('mcq', '') == '')
                    if absent:
                        if not is_optional:
                            has_absent = True
                        cq = mcq = prac = tot = 0
                        lg, gp = 'Ab', 0.0
                        sub_passed = True  # absent — does not count as component-level fail
                    else:
                        cq     = min(int(m.get('cq')  or 0), sub['cqMax'])
                        mcq    = min(int(m.get('mcq') or 0), sub['mcqMax'])
                        prac   = min(int(m.get('prac') or 0), 25) if sub['hasPrac'] else 0
                        theory = cq + mcq
                        tot    = theory + prac
                        lg, gp = _grade_letter(tot)
                        # Check component-level pass marks (applies to optional subjects too,
                        # but optional subject failure does NOT cause overall failure)
                        sub_passed = _subject_passed(
                            cq, mcq, prac,
                            bool(sub.get('hasPrac')),
                            sub.get('cqMax', 70),
                            sub.get('mcqMax', 30),
                        )
                        if not is_optional:
                            if lg == 'F' or not sub_passed:
                                has_fail = True

                    if not absent:
                        if is_optional:
                            if gp > 2.0:
                                total_gpa += (gp - 2.0)
                        else:
                            total_gpa += gp
                            cnt       += 1
                            if lg == 'F' or not sub_passed:
                                fail_cnt += 1   # F grade OR component-level fail
                        total_mark += tot

                    row += ['Ab' if absent else cq,
                            'Ab' if absent else mcq,
                            ('Ab' if absent else prac) if sub['hasPrac'] else '']

                avg    = 0.0 if (fail_cnt > 0 or has_absent or has_fail) else (min(round(total_gpa / cnt, 2), 5.0) if cnt else 0.0)
                passed = fail_cnt == 0 and cnt > 0 and not has_absent and not has_fail
                row   += [total_mark or '', avg if cnt else '',
                          ('Pass' if passed else 'Fail') if cnt else 'Ab']
                writer.writerow(row)
            writer.writerow([])

    output.seek(0)
    return Response(
        '\ufeff' + output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition':
                 f'attachment; filename=Results_{group_filter or "All"}_{cls_filter or "AllClasses"}.csv'}
    )


# ─────────────────────────────────────────────
# ANALYTICS — Promotion, Archive, Roll Gen, Detain, TC
# ─────────────────────────────────────────────

def _compute_student_result(sid, marks_data, group, optional_subject='', cls=None):
    """Return (total_marks, gpa, passed) for a student given their marks dict.
    
    Uses optional_subject (e.g. '178/179') from the student record to determine
    which optional subject pair is active. Falls back to selectedOptional in marks
    for backwards compatibility.
    """
    all_exams = marks_data.get(sid, {})
    if not all_exams:
        return 0, 0.0, False

    exam_key  = list(all_exams.keys())[-1]
    stu_marks = all_exams[exam_key]

    if not cls or not optional_subject:
        stu = Student.query.filter_by(id=sid).first()
        if stu:
            if not cls:
                cls = stu.cls
            if not optional_subject:
                optional_subject = getattr(stu, 'optional_subjects', '') or ''
        if not cls:
            cls = 'Class-XI'

    # Resolve subjects
    subs = _resolve_optional_subjects(group, optional_subject)

    # Filter subjects based on student's class (Class-XI -> 1st Paper; Class-XII -> 2nd Paper)
    if cls == 'Class-XI':
        subs = [sub for sub in subs if not ('2nd' in sub['name'] or '2nd Paper' in sub['name'] or sub['name'].endswith(' 2nd'))]
    elif cls == 'Class-XII':
        subs = [sub for sub in subs if not ('1st' in sub['name'] or '1st Paper' in sub['name'] or sub['name'].endswith(' 1st'))]

    # Dynamically rename Sociology (116) and Islamic History & Culture (267) to append paper number
    import copy
    temp_subs = []
    for sub in subs:
        s = copy.deepcopy(sub)
        if s['code'] in ['116', '267']:
            suffix = ' 1st Paper' if cls == 'Class-XI' else ' 2nd Paper'
            s['name'] = s['name'].replace(' 1st Paper', '').replace(' 2nd Paper', '') + suffix
        temp_subs.append(s)
    subs = temp_subs

    # Backwards-compat: selectedOptional from marks entry (single code)
    selected_optional = stu_marks.get('selectedOptional', '')

    # Resolve optional codes
    opt_codes = [c.strip() for c in (optional_subject or '').split('/') if c.strip()]
    if not opt_codes and selected_optional:
        LEGACY_MAP = {
            'scienceBio': ['178', '179'],
            'scienceMath': ['265', '266'],
            'humLogic': ['121', '122'],
            'humHome': ['273', '274'],
            'busEcon': ['109', '110'],
            'busHome': ['273', '274'],
        }
        opt_codes = LEGACY_MAP.get(selected_optional, [])

    total_gpa = cnt = fail_cnt = total_mark = 0
    has_absent = False
    has_fail = False

    for sub in subs:
        is_optional = sub.get('optional', False) and (sub['code'] in opt_codes)
        m = stu_marks.get(sub['code'], {})
        # Absent = no mark row, explicit absent flag, or both cq/mcq empty strings
        absent = (not m) or m.get('absent', False) or \
                 (str(m.get('cq', '')) == '' and str(m.get('mcq', '')) == '')

        if absent:
            if not is_optional:
                has_absent = True
            continue

        cq     = min(int(m.get('cq')  or 0), sub.get('cqMax', 70))
        mcq    = min(int(m.get('mcq') or 0), sub.get('mcqMax', 30))
        prac   = min(int(m.get('prac') or 0), 25) if sub.get('hasPrac') else 0
        theory = cq + mcq
        tot    = theory + prac
        lg, gp = _grade_letter(tot)
        # Check component-level pass marks for EVERY subject (including optional).
        # Optional subject component-level failure does NOT cause overall student failure.
        sub_passed = _subject_passed(
            cq, mcq, prac,
            bool(sub.get('hasPrac')),
            sub.get('cqMax', 70),
            sub.get('mcqMax', 30),
        )
        if not is_optional:
            if lg == 'F' or not sub_passed:
                has_fail = True

        if is_optional:
            # 4th subject rule: only GP above 2.0 contributes to GPA
            if gp > 2.0:
                total_gpa += (gp - 2.0)
            # Optional subject failure (including component-level) does NOT fail the student
        else:
            total_gpa += gp
            cnt       += 1
            if lg == 'F' or not sub_passed:
                fail_cnt += 1   # F grade OR component-level fail causes fail

        total_mark += tot

    avg = 0.0 if (fail_cnt > 0 or has_absent or has_fail) else (min(round(total_gpa / cnt, 2), 5.0) if cnt else 0.0)
    # passed: no failures AND appeared in at least 1 subject
    return total_mark, avg, (fail_cnt == 0 and cnt > 0 and not has_absent and not has_fail)


@app.route('/api/analyze-promotion', methods=['GET'])
@require_auth
def analyze_promotion():
    """Return Class-XI students eligible for promotion (passed in marks)."""
    students = Student.query.filter_by(cls='Class-XI').all()
    marks    = _get_marks_dict()
    eligible = []

    for stu in students:
        total, gpa, passed = _compute_student_result(stu.id, marks, stu.group, getattr(stu, "optional_subjects", "") or "")
        data = stu.to_dict()
        data.update({
            'total_marks': total,
            'gpa':         gpa,
            'result':      'Pass' if passed else 'Fail',
            'eligible':    passed,
        })
        eligible.append(data)

    eligible.sort(key=lambda x: (-x['gpa'], -x['total_marks']))
    return jsonify({'ok': True, 'data': eligible,
                    'summary': {
                        'total':    len(eligible),
                        'eligible': sum(1 for e in eligible if e['eligible']),
                        'failed':   sum(1 for e in eligible if not e['eligible']),
                    }})


@app.route('/api/execute-promotion', methods=['POST'])
@require_auth
def execute_promotion():
    """Promote eligible Class-XI students to Class-XII."""
    body        = request.get_json(force=True, silent=True) or {}
    student_ids = body.get('studentIds', [])

    if not student_ids:
        return jsonify({'ok': False, 'message': 'No student IDs provided'}), 400

    students    = Student.query.all()
    marks       = _get_marks_dict()
    to_promote  = [s for s in students if s.id in student_ids and s.cls == 'Class-XI']

    if not to_promote:
        return jsonify({'ok': False, 'message': 'No eligible Class-XI students found'}), 400

    scored = []
    for stu in to_promote:
        total, gpa, _ = _compute_student_result(stu.id, marks, stu.group, getattr(stu, "optional_subjects", "") or "")
        scored.append((stu, gpa, total))
    scored.sort(key=lambda x: (-x[1], -x[2]))

    existing_xii_rolls = set()
    for s in students:
        if s.cls == 'Class-XII':
            try:
                existing_xii_rolls.add(int(s.roll))
            except (ValueError, TypeError):
                pass

    next_roll = 1
    promoted_count = 0
    roll_assignments = []

    for stu, gpa, total in scored:
        while next_roll in existing_xii_rolls:
            next_roll += 1
        roll_assignments.append((stu.id, str(next_roll), gpa, total))
        existing_xii_rolls.add(next_roll)
        next_roll += 1

    student_map = {s.id: s for s in students}
    for sid, new_roll, gpa, total in roll_assignments:
        stu = student_map.get(sid)
        if stu:
            old_roll  = stu.roll
            stu.cls   = 'Class-XII'
            stu.roll  = new_roll
            db.session.add(stu)
            db.session.add(PromotionLog(
                student_id=sid,
                name=stu.name,
                old_roll=old_roll,
                new_roll=new_roll,
                gpa=gpa,
                total_marks=total,
            ))
            promoted_count += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Database error during promotion: {str(e)}'}), 500
    return jsonify({'ok': True, 'promoted': promoted_count,
                    'message': f'Successfully promoted {promoted_count} students to Class-XII.'})


@app.route('/api/archive-graduates', methods=['POST'])
@require_auth
def archive_graduates():
    """Move passing Class-XII students to the archive."""
    students = Student.query.filter_by(cls='Class-XII').all()
    marks    = _get_marks_dict()

    to_archive = []
    for stu in students:
        total, gpa, passed = _compute_student_result(stu.id, marks, stu.group, getattr(stu, "optional_subjects", "") or "")
        if passed and total > 0:
            db.session.add(Archive(
                id=stu.id, name=stu.name, roll=stu.roll, reg=stu.reg,
                cls=stu.cls, group=stu.group, section=stu.section,
                father=stu.father, mother=stu.mother, dob=stu.dob,
                phone=stu.phone, religion=stu.religion, year=stu.year,
                session=stu.session, photo=stu.photo,
                total_marks=total, gpa=gpa,
            ))
            to_archive.append(stu)

    if not to_archive:
        return jsonify({'ok': True, 'archived': 0,
                        'message': 'No passing Class-XII graduates found to archive.'})

    for stu in to_archive:
        db.session.delete(stu)

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'message': f'Database error during archiving: {str(e)}'}), 500
    return jsonify({'ok': True, 'archived': len(to_archive),
                    'message': f'{len(to_archive)} graduates archived successfully.'})


@app.route('/api/archive', methods=['GET'])
@require_auth
def get_archive():
    q     = (request.args.get('q') or '').lower()
    query = Archive.query

    if q:
        query = query.filter(
            (Archive.name.ilike(f'%{q}%')) |
            (Archive.roll.ilike(f'%{q}%'))
        )

    return jsonify({'ok': True, 'data': [a.to_dict() for a in query.all()]})


@app.route('/api/analyze-archive-candidates', methods=['GET'])
@require_auth
def analyze_archive_candidates():
    """Returns Class-XII students who have passed and can be moved to archive."""
    students   = Student.query.filter_by(cls='Class-XII').all()
    marks      = _get_marks_dict()
    candidates = []

    for stu in students:
        total, gpa, passed = _compute_student_result(stu.id, marks, stu.group, getattr(stu, "optional_subjects", "") or "")
        if passed:
            data = stu.to_dict()
            data.update({'totalMarks': total, 'gpa': gpa, 'result': 'Pass'})
            candidates.append(data)

    return jsonify({'ok': True, 'data': candidates, 'count': len(candidates)})


@app.route('/api/generate-rolls', methods=['POST'])
@require_auth
def generate_rolls():
    """Sort Class-XII students by GPA and return proposed roll assignments."""
    body         = request.get_json(force=True, silent=True) or {}
    group_filter = body.get('group', '').strip()

    students = Student.query.filter_by(cls='Class-XII').all()
    if group_filter:
        students = [s for s in students if s.group == group_filter]

    marks  = _get_marks_dict()
    scored = []
    for stu in students:
        total, gpa, _ = _compute_student_result(stu.id, marks, stu.group, getattr(stu, "optional_subjects", "") or "")
        data = stu.to_dict()
        data.update({'gpa': gpa, 'totalMarks': total})
        scored.append(data)

    scored.sort(key=lambda x: (-x['gpa'], -x['totalMarks']))

    assignments = [
        {
            'id':         stu['id'],
            'name':       stu['name'],
            'oldRoll':    stu['roll'],
            'newRoll':    str(i),
            'group':      stu.get('group', ''),
            'gpa':        stu['gpa'],
            'totalMarks': stu['totalMarks'],
        }
        for i, stu in enumerate(scored, 1)
    ]

    return jsonify({'ok': True, 'data': assignments, 'count': len(assignments)})


@app.route('/api/detain-list', methods=['GET'])
@require_auth
def detain_list():
    """Return Class-XI students who have failed."""
    students = Student.query.filter_by(cls='Class-XI').all()
    marks    = _get_marks_dict()
    detained = []

    for stu in students:
        total, gpa, passed = _compute_student_result(stu.id, marks, stu.group, getattr(stu, "optional_subjects", "") or "")
        if not passed and total > 0:
            data = stu.to_dict()
            data.update({'total_marks': total, 'gpa': gpa, 'result': 'Fail'})
            detained.append(data)

    return jsonify({'ok': True, 'data': detained})


@app.route('/api/re-enroll/<sid>', methods=['POST'])
@require_auth
def re_enroll(sid):
    """Re-enroll a detained student — reset year/session."""
    body    = request.get_json(force=True, silent=True) or {}
    student = Student.query.filter_by(id=sid).first()

    if not student:
        return jsonify({'ok': False, 'message': 'Student not found'}), 404

    new_year       = body.get('year', datetime.utcnow().strftime('%Y'))
    student.year   = new_year
    student.session = new_year
    student.cls    = 'Class-XI'

    db.session.commit()
    return jsonify({'ok': True, 'message': f'Student re-enrolled for {new_year}.',
                    'data': student.to_dict()})


@app.route('/api/generate-certificate/<sid>', methods=['GET'])
@require_auth
def generate_certificate(sid):
    """Return student data formatted for TC/Certificate generation."""
    stu = Student.query.filter_by(id=sid).first()

    # Also check archive
    if not stu:
        stu = Archive.query.filter_by(id=sid).first()

    if not stu:
        return jsonify({'ok': False, 'message': 'Student not found'}), 404

    marks           = _get_marks_dict()
    total, gpa, passed = _compute_student_result(sid, marks, getattr(stu, "group", "Science"), getattr(stu, "optional_subjects", "") or "")

    settings = {s.key: s.value for s in Setting.query.all()}

    cert_data = {
        'student': stu.to_dict(),
        'college': {
            'name':    settings.get('collegeName', 'Moinuddin Adarsha Mohila College'),
            'address': settings.get('address', 'Sylhet, Bangladesh'),
            'eiin':    settings.get('eiin', ''),
            'phone':   settings.get('phone', ''),
        },
        'result': {
            'total_marks': total,
            'gpa':         gpa,
            'passed':      passed,
            'issue_date':  datetime.utcnow().strftime('%d %B %Y'),
        }
    }
    return jsonify({'ok': True, 'data': cert_data})





# ─────────────────────────────────────────────
# Error handlers
# ─────────────────────────────────────────────
@app.errorhandler(404)
def not_found(e):
    return jsonify({'ok': False, 'message': str(e)}), 404

@app.errorhandler(405)
def method_not_allowed(e):
    return jsonify({'ok': False, 'message': 'Method not allowed'}), 405

@app.errorhandler(500)
def server_error(e):
    # Include real error detail so we can diagnose deployment issues.
    # IMPORTANT: Remove the str(e) in production once the bug is confirmed fixed.
    import traceback
    detail = str(e) if str(e) else repr(e)
    return jsonify({'ok': False, 'message': f'Internal server error: {detail}'}), 500


# Run photo migrations: copy base64 from photo to photo_base64, save to file, and replace with URL
with app.app_context():
    try:
        _st_mig = Student.query.filter(Student.photo.like('data:%')).all()
        if _st_mig:
            print(f"[DB Migration] Migrating {len(_st_mig)} student photos to photo_base64...", flush=True)
            for _s in _st_mig:
                try:
                    _s.photo_base64 = _s.photo
                    _url = _save_photo_file(_s.id, _s.photo)
                    if _url:
                        _s.photo = _url
                    else:
                        _s.photo = ''
                except Exception as _me:
                    print(f"[DB Migration] Student {_s.id} error: {_me}", flush=True)
            db.session.commit()
            print("[DB Migration] Student photo migration done!", flush=True)
    except Exception as _e:
        print(f"[DB Migration] Student photo migration failed: {_e}", flush=True)
        try:
            db.session.rollback()
        except Exception:
            pass

    try:
        _ar_mig = Archive.query.filter(Archive.photo.like('data:%')).all()
        if _ar_mig:
            print(f"[DB Migration] Migrating {len(_ar_mig)} archived photos to photo_base64...", flush=True)
            for _a in _ar_mig:
                try:
                    _a.photo_base64 = _a.photo
                    _url = _save_photo_file(_a.id, _a.photo)
                    if _url:
                        _a.photo = _url
                    else:
                        _a.photo = ''
                except Exception as _me:
                    print(f"[DB Migration] Archive {_a.id} error: {_me}", flush=True)
            db.session.commit()
            print("[DB Migration] Archive photo migration done!", flush=True)
    except Exception as _e:
        print(f"[DB Migration] Archive photo migration failed: {_e}", flush=True)
        try:
            db.session.rollback()
        except Exception:
            pass


# ─────────────────────────────────────────────
# Run
# ─────────────────────────────────────────────
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        print("=" * 60)
        print("  HSC Academic Management System — Flask Backend")
        print("  Moinuddin Adarsha Mohila College, Sylhet")
        print("  PostgreSQL Database Edition")
        print("=" * 60)
        db_url = _db_url
        # Mask password in log output
        safe_url = db_url.split('@')[-1] if '@' in db_url else db_url
        print(f"  Database: ...@{safe_url}")
        print(f"  Photos dir: {PHOTOS_DIR}")
        print()
        print("  Open in browser: http://localhost:5000")
        print("  API Health:      http://localhost:5000/api/health")
        print("=" * 60)

    app.run(debug=True, host='0.0.0.0', port=5000)
