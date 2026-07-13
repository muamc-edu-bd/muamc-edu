"""
Generate a student import template Excel file.
Run: python student_import_template.py

NOTE: The 'Photo' column has been intentionally removed from bulk import.
Photos must be uploaded individually per student via the web interface.
This prevents the JSON bloat bug where base64 images were stored in students.json.
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    # Photo column intentionally excluded — upload photos individually via UI
    headers = ['Name', 'Roll', 'Registration', 'Class', 'Group', 'Section',
               'Father', 'Mother', 'DOB', 'Phone', 'Religion', 'Year', 'Optional Subjects']

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(name='Arial', bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0D2B45", end_color="0D2B45", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Sample data rows
    sample_data = [
        ['Fatima Begum',  '2024001', 'REG001', 'Class-XI', 'Science',    'A', 'Ahmed Ali',      'Rahela Begum',  '2008-01-15', '01712345678', 'Islam', '2024-2025', '178/179'],
        ['Sumaiya Khanam','2024002', 'REG002', 'Class-XI', 'Humanities', 'B', 'Karim Uddin',    'Nasrin Akter',  '2008-03-22', '01798765432', 'Islam', '2024-2025', '109/110'],
        ['Nusrat Jahan',  '2024003', 'REG003', 'Class-XI', 'Business',   'A', 'Abdul Rahman',   'Moriam Begum',  '2008-05-10', '01856789012', 'Islam', '2024-2025', '292/293'],
    ]

    col_widths = [18, 12, 14, 12, 14, 10, 16, 16, 12, 14, 12, 12, 14]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = Font(name='Arial')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # ── Instructions sheet ──
    ws_inst = wb.create_sheet("Instructions")
    ws_inst['A1'] = "Student Bulk Import — Instructions"
    ws_inst['A1'].font = Font(name='Arial', bold=True, size=14, color="0D2B45")

    instructions = [
        "",
        "Column Requirements:",
        "  • Name (required)         — Student's full name",
        "  • Roll (required)         — Roll number; must be unique",
        "  • Registration            — Board registration number",
        "  • Class                   — Class-XI  or  Class-XII",
        "  • Group                   — Science, Humanities, or Business",
        "  • Section                 — A / B / C etc.",
        "  • Father                  — Father's full name",
        "  • Mother                  — Mother's full name",
        "  • DOB                     — Date of birth (YYYY-MM-DD)",
        "  • Phone                   — Contact number",
        "  • Religion                — Islam / Hindu / Christian / Other",
        "  • Year                    — Academic year e.g. 2024-2025",
        "  • Optional Subjects       — Subject codes e.g. 178/179",
        "",
        "Rules:",
        "  • Name and Roll are required; all other fields are optional.",
        "  • Duplicate roll numbers will be skipped automatically.",
        "  • Default Class  : Class-XI   (if not provided)",
        "  • Default Group  : Science    (if not provided)",
        "  • Date format    : YYYY-MM-DD (e.g. 2008-06-15)",
        "  • Save the file as .xlsx before uploading.",
        "",
        "Photos:",
        "  Photos are NOT included in the bulk import to prevent",
        "  database bloat.  Upload each student's photo individually",
        "  using the Edit button in the Student List tab.",
        "",
        "How to Import:",
        "  1. Go to Student Registration tab.",
        "  2. Click 'Bulk Import from Excel'.",
        "  3. Choose your filled Excel file.",
        "  4. Review the preview that appears.",
        "  5. Click 'Import Students' to complete.",
    ]

    for row_idx, line in enumerate(instructions, 2):
        ws_inst[f'A{row_idx}'] = line
        ws_inst[f'A{row_idx}'].font = Font(name='Arial')
        if line.strip().startswith('•'):
            ws_inst[f'A{row_idx}'].alignment = Alignment(wrap_text=True, indent=2)

    ws_inst.column_dimensions['A'].width = 72

    wb.save('student_import_template.xlsx')
    print("✓ Template created: student_import_template.xlsx")
    print("  Note: Photo column removed — upload photos individually via the web UI.")

except ImportError:
    print("❌ openpyxl not installed. Run: pip install openpyxl")
